import functools
import json
import logging
import os
import pathlib
import shutil
import tempfile
from collections import defaultdict
from datetime import datetime
from itertools import groupby
from pathlib import PurePath
from urllib.parse import quote, urlunparse
from uuid import uuid4

import rtyaml
import structlog
import trestle.oscal.component as trestlecomponent
import trestle.oscal.ssp as trestlessp
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import DatabaseError
from django.db.models import Count, Q
from django.db.models.functions import Lower
from django.http import (
    Http404,
    HttpResponse,
    HttpResponseForbidden,
    HttpResponseNotAllowed,
    HttpResponseRedirect,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.text import slugify
from django.views import View
from django.views.generic import ListView
from natsort import natsorted
from simple_history.utils import update_change_reason
from structlog import get_logger
from structlog.stdlib import LoggerFactory

import controls.utils as utils
import tools.diff_match_patch.python3 as dmp_module
from controls.oscal import Catalog, CatalogData, Catalogs
from siteapp.models import Invitation, Organization, Project, Tag
from siteapp.settings import GOVREADY_URL
from system_settings.models import SystemSettings
from utils import project_navigation

from .forms import (
    DeploymentForm,
    ElementEditForm,
    ElementForm,
    ImportOSCALComponentForm,
    PoamForm,
    StatementPoamForm,
    SystemAssessmentResultForm,
)
from .models import (
    Deployment,
    Element,
    ElementControl,
    ImportRecord,
    Poam,
    Statement,
    StatementTypeEnum,
    System,
    SystemAssessmentResult,
)
from .utilities import (
    get_control_statement_part,
    increment_element_name,
    oscalize_catalog_key,
    oscalize_control_id,
)

logging.basicConfig()

structlog.configure(logger_factory=LoggerFactory())
structlog.configure(processors=[structlog.processors.JSONRenderer()])
logger = get_logger()


def index(request):
    """Index page for controls"""

    # Get catalog
    catalog = Catalog()
    control_groups = catalog.get_groups()
    context = {
        "catalog": catalog,
        "control": None,
        "common_controls": None,
        "control_groups": control_groups,
    }
    return render(request, "controls/index.html", context)


def catalogs(request):
    """Index page for catalogs"""

    context = {"catalogs": Catalogs.catalogs()}
    return render(request, "controls/index-catalogs.html", context)


def catalog(request, catalog_key, system_id=None):
    """Index page for controls"""

    if system_id is None:
        system = None
    else:
        system = System.objects.get(pk=system_id)

    # Get catalog
    catalog = Catalog(catalog_key)
    control_groups = catalog.get_groups()
    context = {
        "catalog": catalog,
        "control": None,
        "common_controls": None,
        "system": system,
        "control_groups": control_groups,
    }
    return render(request, "controls/index.html", context)


def group(request, catalog_key, g_id):
    """Temporary index page for catalog control group"""

    # Get catalog
    catalog = Catalog(catalog_key)
    control_groups = catalog.get_groups()
    group = None
    # Get group/family of controls
    for g in control_groups:
        if g["id"].lower() == g_id:
            group = g
            break

    context = {
        "catalog": catalog,
        "control": None,
        "common_controls": None,
        "control_groups": control_groups,
        "group": group,
    }
    return render(request, "controls/index-group.html", context)


def control(request, catalog_key, cl_id):
    """Control detail view"""
    cl_id = oscalize_control_id(cl_id)
    catalog_key = oscalize_catalog_key(catalog_key)

    # Get catalog
    catalog = Catalog(catalog_key)
    cg_flat = catalog.get_flattened_controls_all_as_dict()
    # Prepare links
    links = []
    for link in cg_flat[cl_id.lower()]["guidance_links"]:
        link["href_split"] = link["href"].split("/")
        if len(link["href_split"]) == 6:
            link["catalog"] = link["href_split"][3].replace("_", " ")
        else:
            link["catalog"] = None
        links.append(link)

    # Handle properly formatted control id that does not exist
    if cl_id.lower() not in cg_flat:
        return render(
            request, "controls/detail.html", {"catalog": catalog, "control": {}}
        )
    # Get and return the control
    context = {
        "catalog": catalog,
        "control": cg_flat[cl_id.lower()],
        "links": links,
    }
    return render(request, "controls/detail.html", context)


def controls_selected(request, system_id):
    """Display System's selected controls view"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("view_system", system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project

        project = system.projects.all()[0]
        controls = system.root_element.controls.all()

        # sort controls
        controls = list(controls)
        try:
            controls.sort(
                key=lambda control: control.get_flattened_oscal_control_as_dict()[
                    "sort_id"
                ]
            )
        except Exception as e:
            logger.error(f"Error flattening controls: {e}")

        paginator = Paginator(controls, 25)
        page_number = request.GET.get("page", 1)
        list_items = paginator.get_page(page_number)
        pager = paginator.get_elided_page_range(
            number=page_number, on_each_side=1, on_ends=1
        )

        # Determine if a legacy statement exists for the control
        impl_smts_legacy = Statement.objects.filter(
            consumer_element=system.root_element,
            statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_LEGACY.name,
        )
        impl_smts_legacy_dict = {}
        for legacy_smt in impl_smts_legacy:
            impl_smts_legacy_dict[legacy_smt.sid] = legacy_smt

        # collect component counts by statement id
        component_counts = (
            system.root_element.statements_consumed.filter(
                statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION.name
            )
            .values("sid")
            .annotate(component_count=Count("producer_element_id"))
            .order_by("sid")
        )
        impl_smts_cmpts_count = dict(
            (c_count["sid"], c_count["component_count"]) for c_count in component_counts
        )

        # Get list of catalog objects

        # TODO: we should fix the template so that we don't require
        # this magic.
        # Template has hardcoded links for some catalogs, so we remove
        # them from the list.

        internal_catalog_keys = [
            Catalogs.CMMC_ver1,
            Catalogs.NIST_SP_800_171_rev1,
            Catalogs.NIST_SP_800_53_rev4,
            Catalogs.NIST_SP_800_53_rev5,
        ]
        external_catalogs = [
            catalog
            for catalog in Catalogs.catalogs()
            if catalog.catalog_key not in internal_catalog_keys
        ]

        nav = project_navigation(request, project)

        # Return the controls
        context = {
            "system": system,
            "project": project,
            "list_items": list_items,
            "pager": pager,
            "external_catalogs": external_catalogs,
            "impl_smts_cmpts_count": impl_smts_cmpts_count,
            "impl_smts_legacy_dict": impl_smts_legacy_dict,
            "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
            "security_sensitivity": utils.get_security_sensitivity(project),
            "send_invitation": Invitation.form_context_dict(
                request.user, project, [request.user]
            ),
            "nav": nav,
        }

        html = render(request, "systems/controls_selected.html", context)
        return html
    else:
        # User does not have permission to this system
        raise Http404


@login_required
def system_controls_add(request, system_id):
    """Add a selected control to a system (e.g., selected controls)"""

    # Get control values from request.POST
    catalog_key = request.POST["catalog_key"].replace(
        " ", "_"
    )  # Make sure catalog key has underscores instead of spaces
    control_id = request.POST["control_id"]

    system = System.objects.get(id=system_id)
    controls = system.root_element.controls.all()

    if controls.filter(Q(oscal_catalog_key=catalog_key)).filter(
        Q(oscal_ctl_id=control_id)
    ):
        messages.add_message(
            request,
            messages.WARNING,
            f"Control {control_id.upper()} in catalog {catalog_key} can't be added because \
                it is already included in your system controls.",
        )
        # Log result
        logger.warning(
            event="change_system add_selected_control",
            object={"object": "control", "id": control_id, "catalog": catalog_key},
            user={"id": request.user.id, "username": request.user.username},
        )
        return redirect(reverse("controls_selected", args=[system_id]))

    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("change_system", system):

        # Add ElementControl to system
        system.add_control(catalog_key, control_id)

        # Create message for user
        messages.add_message(
            request,
            messages.INFO,
            f"Control {control_id.upper()} was added to selected controls.",
        )

    else:
        # User does not have permission
        # Log result
        logger.info(
            event="change_system add_selected_control permission_denied",
            object={"object": "control", "id": control_id},
            user={"id": request.user.id, "username": request.user.username},
        )

        # Create message for user
        messages.add_message(
            request, messages.INFO, "You don't have permission to edit the system."
        )

    response = redirect(reverse("controls_selected", args=[system_id]))
    return response


@login_required
def system_control_remove(request, system_id, element_control_id):
    """Remove a selected control from a system and delete/hide the related statements"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("change_system", system):

        # Retrieve ElementControl
        ec = ElementControl.objects.get(id=element_control_id)

        # Delete the control implementation statements associated with this component
        system.remove_control(element_control_id)
        messages.add_message(
            request, messages.INFO, f"Control {ec.oscal_ctl_id} was removed."
        )

        # Log result
        logger.info(
            event="change_system remove_selected_control",
            object={"object": "control", "id": element_control_id},
            user={"id": request.user.id, "username": request.user.username},
        )

    else:
        # User does not have permission
        # Log result
        logger.info(
            event="change_system remove_selected_control permission_denied",
            object={"object": "control", "id": element_control_id},
            user={"id": request.user.id, "username": request.user.username},
        )

        # Create message for user
        messages.add_message(
            request, messages.INFO, "You don't have permission to edit the system."
        )

    response = redirect(reverse("controls_selected", args=[system_id]))
    return response


@functools.lru_cache()
def controls_updated(request, system_id):
    """Display System's statements by updated date in reverse chronological order"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("view_system", system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        controls = system.root_element.controls.all()

        impl_smts_count = {}
        ikeys = system.smts_control_implementation_as_dict.keys()
        for c in controls:
            impl_smts_count[c.oscal_ctl_id] = 0
            if c.oscal_ctl_id in ikeys:
                impl_smts_count[c.oscal_ctl_id] = len(
                    system.smts_control_implementation_as_dict[c.oscal_ctl_id][
                        "control_impl_smts"
                    ]
                )

        # Return the controls
        context = {
            "system": system,
            "project": project,
            "controls": controls,
            "impl_smts_count": impl_smts_count,
            "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
        }
        return render(request, "systems/controls_updated.html", context)
    else:
        # User does not have permission to this system
        raise Http404


@login_required
def edit_element(request, element_id):
    """
      Edit Element information as long as the name does not clash with a different element name
    Args:
        request ([HttpRequest]): The network request
    component_id ([int|str]): The id of the component
    Returns:
        [JsonResponse]: Either a ok status or an error
    """

    # The original element(component)
    ele_instance = get_object_or_404(Element, id=element_id)

    if request.method == "POST":
        new_name = request.POST.get("name", "").strip() or None

        if (
            Element.objects.filter(name__iexact=new_name).exists()
            and new_name != ele_instance.name
        ):
            return JsonResponse({"status": "err", "message": "Name already in use"})

        form = ElementEditForm(request.POST or None, instance=ele_instance)
        if form.is_valid():
            logger.info(
                event="edit_element",
                object={
                    "object": "element",
                    "id": form.instance.id,
                    "name": form.instance.name,
                },
                user={"id": request.user.id, "username": request.user.username},
            )
            form.save()
            return JsonResponse({"status": "ok"})
        else:
            errors = form.errors.get_json_data(escape_html=False)
            msg_list = [
                f"{e.title()} - {errors[e][0]['message']}" for e in errors.keys()
            ]
            return JsonResponse(
                {
                    "status": "err",
                    "message": "Please fix the following problems:<br>"
                    + "<br>".join(msg_list),
                }
            )


class SelectedComponentsList(ListView):
    """
    Display System's selected components view
    """

    model = Element
    template_name = "systems/components_selected.html"
    context_object_name = "system_elements"
    ordering = ["name"]
    paginate_by = 25

    def get_queryset(self):
        """
        Return the systems producer elements.
        """
        system = System.objects.get(id=self.kwargs["system_id"])
        return [
            element
            for element in system.producer_elements
            if element.element_type != "system"
        ]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Retrieve identified System
        system = System.objects.get(id=self.kwargs["system_id"])

        # Retrieve related selected controls if user has permission on system
        if self.request.user.has_perm("view_system", system):
            # Retrieve primary system Project
            # Temporarily assume only one project and get first project
            project = system.projects.first()
            context["project"] = project
            context["system"] = system
            context["elements"] = Element.objects.all().exclude(element_type="system")
            context["nav"] = project_navigation(self.request, project)
            context["send_invitation"] = (
                Invitation.form_context_dict(
                    self.request.user, project, [self.request.user]
                ),
            )
            return context
        else:
            # User does not have permission to this system
            raise Http404


@login_required
def component_library(request):
    """Display the library of components"""

    query = Q()
    keyword = request.GET.get("search")
    if keyword:
        query &= Q(name__icontains=keyword)
        query |= Q(description__icontains=keyword)
    types = request.GET.get("type")
    if types:
        t = types.split(",")
        query &= Q(component_type__in=t)

    try:
        element_list = (
            Element.objects.filter(query).exclude(element_type="system").distinct()
        )
    except DatabaseError as e:
        logger.error(f"A database error occured: {e}")

    # Natural sorting on name
    element_list = natsorted(element_list, key=lambda x: x.name)

    # Pagination
    paginator = Paginator(element_list, 15)
    page_number = request.GET.get("page", 1)
    pager = None

    try:
        list_items = paginator.page(page_number)
    except PageNotAnInteger:
        list_items = paginator.page(1)
    except EmptyPage:
        list_items = paginator.page(paginator.num_pages)

    if list_items.paginator.num_pages > 0:
        pager = paginator.get_elided_page_range(
            number=page_number, on_each_side=1, on_ends=1
        )

    filters = {}
    types = (
        Element.objects.all()
        .values("component_type")
        .filter(component_type__isnull=False)
        .exclude(element_type="system")
        .order_by("component_type")
        .annotate(Count("component_type"))
    )
    if types:
        filters["type"] = []
        for t in types:
            filters["type"].append(
                {
                    "name": t.get("component_type"),
                    "count": t.get("component_type__count"),
                }
            )

    context = {
        "list_items": list_items,
        "pager": pager,
        "import_form": ImportOSCALComponentForm(),
        "total_comps": Element.objects.exclude(element_type="system").count(),
        "filters": filters,
    }

    return render(request, "components/component_library.html", context)


def diff_components_prettyHtml(smt1, smt2):
    """Generate a diff of two statements of type `control_implementation`"""
    dmp = dmp_module.diff_match_patch()
    val1 = ""
    val2 = ""
    if hasattr(smt1, "body"):
        val1 = smt1.body
    if hasattr(smt2, "body"):
        val2 = smt2.body

    diff = dmp.diff_main(val1, val2)
    if len(diff) == 1:
        return "Statement is identical."
    return dmp.diff_prettyHtml(diff)


def compare_components(request):
    """
    Compare submitted components
    """

    checks = json.loads(request.POST.get("hiddenChecks"))
    compare_list = list(checks.values())
    if len(compare_list) <= 1:
        # add messages
        messages.add_message(
            request,
            messages.WARNING,
            "You must select at least two components to compare.",
        )
        return HttpResponseRedirect("/controls/components")
    else:
        ele_q = (
            Element.objects.filter(pk__in=compare_list)
            .exclude(element_type="system")
            .distinct()
        )
        # Maintain sort order of compare_list otherwise Django will order ascending
        element_list = sorted(ele_q, key=lambda x: compare_list.index(str(x.id)))
        compare_prime, element_list = (
            element_list[0],
            element_list[1:],
        )  # The first component selected will be compared against the rest
        compare_prime_smts = compare_prime.statements(
            StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name
        )
    difference_tuples = []
    for component in element_list:
        differences = []
        # compare each component's statements to prime
        cmt_smts = component.statements(
            StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name
        )
        if cmt_smts.exists():
            # TODO: Need to create a tuple with smt id to return appropriate
            for smt in cmt_smts:
                smt_prime = (
                    compare_prime_smts.filter(sid=smt.sid)
                    .filter(pid=smt.pid)
                    .filter(sid_class=smt.sid_class)
                    .first()
                )
                # If it is not a statement found in both components then we just add styling
                if smt_prime:
                    diff = diff_components_prettyHtml(smt_prime, smt)
                else:
                    diff = (
                        f"<span><ins style='background:#e6ffe6;'>{smt.body}</ins><span>"
                    )
                differences.append(diff)
        difference_tuples.extend(
            zip(
                [component.id] * len(cmt_smts),
                [component.name] * len(cmt_smts),
                cmt_smts,
                differences,
            )
        )
    if request.method == "POST":
        context = {
            "element_list": element_list,
            "compare_prime": compare_prime,
            "prime_smts": compare_prime_smts,
            "secondary_smts": cmt_smts,
            "differences": difference_tuples,
            "compare_list": compare_list,
        }
        return render(request, "components/compare_components.html", context)


@login_required
def import_records(request):
    """Display the records of component imports"""

    import_records = ImportRecord.objects.all()
    import_components = {}

    for import_record in import_records:
        import_components[import_record] = Element.objects.filter(
            import_record=import_record
        )

    context = {
        "import_components": import_components,
    }

    return render(request, "components/import_records.html", context)


@login_required
def import_record_details(request, import_record_id):
    """Display the records of component imports"""

    import_record = ImportRecord.objects.get(id=import_record_id)
    component_statements = import_record.get_components_statements()
    statements_prototype = Statement.objects.filter(
        statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name,
        import_record=import_record,
    )

    context = {
        "import_record": import_record,
        "component_statements": component_statements,
        "statements_prototype": statements_prototype,
    }
    return render(request, "components/import_record_details.html", context)


@login_required
def confirm_import_record_delete(request, import_record_id):
    """Delete the components and statements imported from a particular import record"""

    import_record = ImportRecord.objects.get(id=import_record_id)
    component_statements = import_record.get_components_statements()
    component_count = len(component_statements)
    statement_count = 0
    for component in component_statements:
        statement_count += component_statements[component].count()
    statements_prototype = Statement.objects.filter(
        statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name,
        import_record=import_record,
    )
    statement_prototype_count = len(statements_prototype)
    projects = import_record.import_record_projects.all()
    project_count = len(projects)
    elements = import_record.import_record_elements.all()
    element_count = len(elements)
    context = {
        "import_record": import_record,
        "component_count": component_count,
        "statement_count": statement_count,
        "statement_prototype_count": statement_prototype_count,
        "project_count": project_count,
        "element_count": element_count,
    }
    return render(request, "components/confirm_import_record_delete.html", context)


@login_required
def import_record_delete(request, import_record_id):
    """Delete the components and statements imported from a particular import record"""

    import_record = ImportRecord.objects.get(id=import_record_id)
    import_created = import_record.created
    import_record.delete()

    messages.add_message(request, messages.INFO, f"{import_created} was deleted.")

    response = redirect("/controls/components")
    return response


class SystemSecurityPlanSerializer(object):
    def __init__(self, system, impl_smts):
        self.system = system
        self.impl_smts = impl_smts


class OSCALSystemSecurityPlanSerializer(SystemSecurityPlanSerializer):
    @staticmethod
    def ssp_statement_id_from_control(control_id, part_id):
        if part_id:
            return f"{control_id}_smt.{part_id}"
        else:
            return f"{control_id}_smt"

    def create_statement_dicts(self, control_id, group_impl_smts):
        """
        Create statements for each group of implementation statements
        """
        statements = []
        for smt in group_impl_smts:
            # Get oscal control and form statement id with part if present
            cl_id = oscalize_control_id(control_id)
            smt_id = self.ssp_statement_id_from_control(cl_id, smt.pid)
            # TODO: "set-parameters" ?
            statement_dict = {
                "statement-id": smt_id,
                "uuid": str(uuid4()),
                "by-components": [
                    {
                        "component-uuid": str(smt.producer_element.uuid),
                        "uuid": str(smt.uuid),
                        "description": smt.body,
                    }
                ],
            }
            statements.append(statement_dict)
        return statements

    def as_json(self):
        project = self.system.projects.first()
        # TODO: again a stub as found in module_logic, this is still not entirely valid.
        profile = urlunparse(
            (GOVREADY_URL.scheme, GOVREADY_URL.netloc, "profile_path", None, None, None)
        )
        orgs = list(
            Organization.objects.filter(projects=project)
        )  # TODO: orgs need uuids
        components = (
            Element.objects.filter(statements_produced__in=self.impl_smts)
            .filter(component_state="operational")
            .exclude(element_type="system")
        )
        impl_comps = [
            {"component-uuid": str(component.uuid)} for component in components
        ]
        parties = [
            {"uuid": str(uuid4()), "type": "organization", "name": org.name}
            for org in orgs
        ]
        of = {
            "system-security-plan": {
                "uuid": str(self.system.root_element.uuid),
                "metadata": {
                    "title": "{} System Security Plan".format(
                        self.system.root_element.name
                    ),
                    "last-modified": self.system.root_element.updated.replace(
                        microsecond=0
                    ).isoformat(),
                    "version": project.version,
                    "oscal-version": self.system.root_element.oscal_version,
                    "roles": [],
                    "parties": parties,
                },
                "import-profile": {"href": profile},
                "system-characteristics": {},
                "system-implementation": {
                    "remarks": "",
                    "users": [],
                    "components": [],
                    "inventory-items": [
                        {
                            "uuid": str(uuid4()),
                            "description": "An inventory item",
                            "implemented-components": impl_comps,
                        }
                    ],  # TODO: inventory-items props, description, responsible-parties
                },
                "control-implementation": {
                    "description": "",
                    "implemented-requirements": [],  # implemented-requirements
                },
            }
        }
        # Create a list of dicts that are the implementation requirements
        # Each element is for each control sid group
        # Each group has that controls statements
        for control_id, group in groupby(
            natsorted(self.impl_smts, key=lambda ismt: ismt.sid), lambda ismt: ismt.sid
        ):
            imp_req_dict = {
                "uuid": str(uuid4()),
                "control-id": "{}".format(control_id),
                "statements": self.create_statement_dicts(control_id, group),
            }  # statements

            of["system-security-plan"]["control-implementation"][
                "implemented-requirements"
            ].append(imp_req_dict)

        # System implementation
        users = project.get_all_participants()
        user_party_uuid = str(uuid4())

        of["system-security-plan"]["metadata"]["roles"] = [
            {
                "id": auths.get("role", "member").split(";")[0],
                "title": auths.get("role", "member").split(";")[0].capitalize(),
            }
            for user, auths in users
        ]
        of["system-security-plan"]["system-implementation"]["users"] = [
            {
                "uuid": user_party_uuid,
                "title": user.username,
                "role-ids": [auths.get("role", "member").split(";")[0]],
            }
            for user, auths in users
        ]
        of["system-security-plan"]["system-implementation"]["components"] = [
            {
                "uuid": str(comp_ele.uuid),
                "title": comp_ele.name,
                "description": comp_ele.description,
                "status": {"state": comp_ele.component_state},
                "type": comp_ele.component_type,
                "responsible-roles": [
                    {"role-id": "asset-owner", "party-uuids": [user_party_uuid]}
                ],
                "props": [
                    {
                        "name": "tag",
                        "ns": "https://govready.com/ns/oscal",
                        "value": tag.label,
                    }
                    for tag in comp_ele.tags.all()
                ],
            }
            for comp_ele in components
        ]  # TODO: responsible-roles

        # System characteristics
        # TODO: status remarks, authorization-boundary
        security_body = project.system.get_security_impact_level
        confidentiality = security_body.get(
            "security_objective_confidentiality", "UNKOWN"
        )
        integrity = security_body.get("security_objective_integrity", "UNKOWN")
        availability = security_body.get("security_objective_availability", "UNKOWN")
        information_types = [
            {
                "uuid": str(uuid4()),
                "title": "UNKNOWN information type title",
                # "categorizations": [], # TODO https://doi.org/10.6028/NIST.SP.800-60v2r1
                "description": "information type description",
                "confidentiality-impact": {"base": confidentiality},
                "integrity-impact": {"base": integrity},
                "availability-impact": {"base": availability},
            }
        ]
        of["system-security-plan"]["system-characteristics"] = {
            "system-name": self.system.root_element.name,
            "description": self.system.root_element.description,
            "system-ids": [
                {
                    "id": str(self.system.root_element.uuid),  # TODO: identifier-type
                    "identifier-type": "https://ietf.org/rfc/rfc4122",
                }
            ],
            "security-sensitivity-level": self.system.get_security_sensitivity_level
            if self.system.get_security_sensitivity_level
            else "UNKOWN",
            "system-information": {"information-types": information_types},
            "security-impact-level": {
                "security-objective-confidentiality": confidentiality,
                "security-objective-integrity": integrity,
                "security-objective-availability": availability,
            },
            "status": {
                "state": self.system.root_element.component_state,
                "remarks": "",
            },
            "authorization-boundary": {
                "description": "The description of the authorization boundary would go here."
            },
        }
        try:
            # Create a temporary directory and dump the json_object in there.
            tempdir = tempfile.mkdtemp()
            path = os.path.join(tempdir, "ssp_object.json")
            # Use trestle's ComponentDefinition method oscal_read to read the path to json in
            # the temporary folder
            path_ssp_definition = pathlib.Path(path)

            with open(path, "w+") as cred:
                json.dump(of, cred)
            # Read in temporary file and shape into trestle pydantic SSP definition.
            trestle_oscal_json = trestlessp.SystemSecurityPlan.oscal_read(
                path_ssp_definition
            )
            # Finally validate that this object is valid by the OSCAL System Security Plan
            # definition
            trestlessp.SystemSecurityPlan.validate(trestle_oscal_json)
            # Cleanup
            shutil.rmtree(tempdir)
        except Exception as e:
            logger.error(f"Invalid System Security Plan JSON: {e}")
            shutil.rmtree(tempdir)
            return HttpResponse(e)

        oscal_string = json.dumps(of, sort_keys=False, indent=2)
        return oscal_string


class ComponentSerializer(object):
    def __init__(self, element, impl_smts):
        self.element = element
        self.impl_smts = impl_smts


class OSCALComponentSerializer(ComponentSerializer):
    @staticmethod
    def statement_id_from_control(control_id, part_id):
        # Checking for a case where the control was provided like ac-2.3 which already has
        # its part included.
        if part_id:
            if part_id not in control_id:
                return f"{control_id}.{part_id}"

        return f"{control_id}"

    def generate_source(self, src_str):
        """Return a valid catalog source given string"""
        DEFAULT_SOURCE = "NIST_SP-800-53_rev5"
        if not src_str:
            return DEFAULT_SOURCE
        # TODO: Handle other cases
        source = src_str
        return source

    def as_json(self):
        # Build OSCAL
        comp_uuid = str(self.element.uuid)
        control_implementations = []
        props = []
        orgs = list(
            Organization.objects.all()
        )  # TODO: orgs need uuids, not sure which orgs to use for a component
        parties = [
            {"uuid": str(uuid4()), "type": "organization", "name": org.name}
            for org in orgs
        ]
        responsible_roles = [
            {
                "role-id": "supplier",  # TODO: Not sure what this refers to
                "party-uuids": [str(party.get("uuid")) for party in parties],
            }
        ]
        of = {
            "component-definition": {
                "uuid": str(uuid4()),
                "metadata": {
                    "title": "{}".format(self.element.name),
                    "last-modified": self.element.updated.replace(
                        microsecond=0
                    ).isoformat(),
                    "version": self.element.updated.replace(microsecond=0).isoformat(),
                    "oscal-version": self.element.oscal_version,
                    "parties": parties,
                },
                "components": [
                    {
                        "uuid": comp_uuid,
                        "type": self.element.component_type.lower()
                        if self.element.component_type is not None
                        else "software",
                        "title": self.element.full_name or self.element.name,
                        "description": self.element.description,
                        "responsible-roles": responsible_roles,
                        "props": props,
                        "control-implementations": control_implementations,
                    }
                ],
            },
        }

        # Add component's tags if they exist
        if self.element.tags.exists():
            props.extend(
                [
                    {
                        "name": "tag",
                        "ns": "https://govready.com/ns/oscal",
                        "value": tag.label,
                    }
                    for tag in self.element.tags.all()
                ]
            )

        # Remove 'metadata.props' key if no metadata.props exist
        if len(props) == 0:
            of["component-definition"]["metadata"].pop("props", None)

        # create requirements and organize by source (sid_class)

        by_class = defaultdict(list)

        # work:
        # group stmts by control-id
        # emit an requirement for the control-id
        # iterate over each group
        # emit a statement for each member of the group
        # notes:
        # - OSCAL implemented_requirements and control_implementations need UUIDs
        #   which we don't have in the db, so we construct them.

        for control_id, group in groupby(
            natsorted(self.impl_smts, key=lambda ismt: ismt.sid), lambda ismt: ismt.sid
        ):

            for smt in group:
                statement_id = self.statement_id_from_control(control_id, smt.pid)
                statement_req = {
                    "uuid": str(smt.uuid),
                    "description": smt.body,
                    "control-id": statement_id,
                }
                # key-value by sid a.k.a control id for each requirement
                if statement_req not in by_class[smt.sid]:
                    by_class[smt.sid].append(statement_req)

        for sid_class, requirements in by_class.items():
            control_implementation = {
                "uuid": str(
                    uuid4()
                ),  # TODO: Not sure if this should implemented or just generated here.
                "source": self.generate_source(smt.source if smt.source else None),
                "description": f"This is a partial implementation of the {sid_class} catalog, \
                     focusing on the control enhancement {requirements[0].get('control-id')}.",
                "implemented-requirements": [req for req in requirements],
            }
            control_implementations.append(control_implementation)
        # Remove 'control-implementations' key if no implementations exist
        if len(control_implementations) == 0:
            of["component-definition"]["components"][0].pop(
                "control-implementations", None
            )

        oscal_string = json.dumps(of, sort_keys=False, indent=2)
        return oscal_string


class OpenControlComponentSerializer(ComponentSerializer):
    def as_yaml(self):
        ocf = {
            "name": self.element.name,
            "schema_version": "3.0.0",
            "documentation_complete": False,
            "satisfies": [],
        }

        satisfies_smts = ocf["satisfies"]
        for smt in self.impl_smts:
            my_dict = {
                "control_key": smt.sid.upper(),
                "control_name": smt.catalog_control_as_dict["title"],
                "standard_key": smt.sid_class,
                "covered_by": [],
                "security_control_type": "Hybrid | Inherited | ...",
                "narrative": [{"text": smt.body}],
                "remarks": [{"text": smt.remarks}],
            }
            satisfies_smts.append(my_dict)
        opencontrol_string = rtyaml.dump(ocf)
        return opencontrol_string


class ComponentImporter(object):
    def import_components_as_json(
        self,
        import_name,
        json_object,
        request=None,
        existing_import_record=False,
        stopinvalid=True,
    ):
        """Imports Components from a JSON object

        @type import_name: str
        @param import_name: Name of import file (if it exists)
        @type json_object: dict
        @param json_object: Element attributes from JSON object
        @type existing_import_record: boolean
        @param existing_import_record: Continue to append imports to an existing import record
        @rtype: ImportRecord if success, bool (false) if failure
        @returns: ImportRecord linked to the created components (if success) or False if
        failure
        """

        issues = []
        try:
            # Create a temporary directory and dump the json_object in there.
            tempdir = tempfile.mkdtemp()
            path = os.path.join(tempdir, "object.json")
            path_component_definition = pathlib.Path(path)
            oscal_json = json.loads(json_object)
            with open(path, "w+") as cred:
                json.dump(oscal_json, cred)
            # Read in temporary file and shape into trestle pydantic Component definition.
            trestle_oscal_json = trestlecomponent.ComponentDefinition.oscal_read(
                path_component_definition
            )
            # Finally validate that this object is valid by the component definition
            trestlecomponent.ComponentDefinition.validate(trestle_oscal_json)
            # Cleanup
            shutil.rmtree(tempdir)
        except Exception as e:
            logger.error(e)
            logger.warning(
                event="error_importing_component",
                object={
                    "object": "component",
                    "name": import_name,
                    "error": {e.__context__},
                },
            )
            issues.append(
                {"object": "component", "name": import_name, "error": {e.__context__}}
            )
            shutil.rmtree(tempdir)
            # Check a component uploaded to form
            if request and request.POST.get("json_content") is not None:
                if stopinvalid:
                    messages.add_message(
                        request,
                        messages.ERROR,
                        f"{e.__context__} is an invalid JSON component. Import stopped.",
                    )
                    return HttpResponse(e)
                else:
                    messages.add_message(
                        request,
                        messages.INFO,
                        f"{e.__context__} is an invalid JSON component. Import continued with \
                            possible error(s).",
                    )
            else:
                if stopinvalid:
                    print("\nNOTICE - ISSUES DURING COMPONENT IMPORT\n")
                    [print(issue) for issue in issues]
                    print("\nPROGRAM HALTED\n")
                    import sys

                    sys.exit()

        # If importing from importcomponents script print issues
        if len(issues) > 0:
            print("\nNOTICE - ISSUES DURING COMPONENT IMPORT\n")
            [print(issue) for issue in issues]

        # Returns list of created components
        created_components = self.create_components(oscal_json)
        new_import_record = self.create_import_record(
            import_name,
            created_components,
            existing_import_record=existing_import_record,
        )
        return new_import_record

    def create_import_record(
        self, import_name, components, existing_import_record=False
    ):
        """Associates components and statements to an import record

        @type import_name: str
        @param import_name: Name of import file (if it exists)
        @type components: list
        @param components: List of components
        @type existing_import_record: booleen
        @param existing_import_record: Continue to append imports to an existing import record
        @rtype: ImportRecord
        @returns: New ImportRecord object with components and statements associated
        """

        import_record = ImportRecord.objects.filter(name=import_name).last()
        if import_record is None or not existing_import_record:
            import_record = ImportRecord.objects.create(name=import_name)
        for component in components:
            statements = Statement.objects.filter(producer_element=component)
            for statement in statements:
                statement.import_record = import_record
                # statement.save()
            component.import_record = import_record
            component.save()

        return import_record

    def create_components(self, oscal_json):
        """Creates Elements (Components) from valid OSCAL JSON"""
        components_created = []
        components = oscal_json["component-definition"]["components"]
        for component in components:
            new_component = self.create_component(component)
            if new_component is not None:
                components_created.append(new_component)

        return components_created

    def create_component(self, component_json):
        """Creates a component from a JSON dict

        @type component_json: dict
        @param component_json: Component attributes from JSON object
        @rtype: Element
        @returns: Element object if created, None otherwise
        """

        component_name = component_json["title"]
        while Element.objects.filter(name=component_name).count() > 0:
            component_name = increment_element_name(component_name)

        new_component = Element.objects.create(
            name=component_name,
            description=component_json["description"]
            if "description" in component_json
            else "Description missing",
            # Components uploaded to the Component Library are all system_element types
            element_type="system_element",
            uuid=component_json["uuid"] if "uuid" in component_json else uuid4(),
            component_type=component_json["type"]
            if "type" in component_json
            else "software",
        )

        logger.info(
            f"Component {new_component.name} created with UUID {new_component.uuid}."
        )

        component_props = component_json.get("props", None)
        if component_props is not None:
            desired_tags = set(
                [
                    prop["value"]
                    for prop in component_props
                    if prop["name"] == "tag"
                    and "ns" in prop
                    and prop["ns"] == "https://govready.com/ns/oscal"
                ]
            )
            existing_tags = Tag.objects.filter(label__in=desired_tags).values(
                "id", "label"
            )
            tags_to_create = desired_tags.difference(
                set([tag["label"] for tag in existing_tags])
            )
            new_tags = Tag.objects.bulk_create(
                [Tag(label=tag) for tag in tags_to_create]
            )
            all_tag_ids = [tag.id for tag in new_tags] + [
                tag["id"] for tag in existing_tags
            ]
            new_component.add_tags(all_tag_ids)
            new_component.save()
        control_implementation_statements = component_json.get(
            "control-implementations", None
        )
        # If there data exists the OSCAL component's control-implementations key
        if control_implementation_statements:
            for control_element in control_implementation_statements:
                catalog = oscalize_catalog_key(control_element.get("source", None))
                created_statements = self.create_control_implementation_statements(
                    catalog, control_element, new_component
                )
        # If there are no valid statements in the json object
        if created_statements == []:
            logger.info(
                f"The Component {new_component.name} will be deleted as there were no valid \
                    statements provided."
            )
            new_component.delete()
            new_component = None

        return new_component

    def create_control_implementation_statements(
        self, catalog_key, control_element, parent_component
    ):
        """Creates a Statement from a JSON dict implemented-requirements

        @type catalog_key: str
        @param catalog_key: Catalog of the control statements
        @type control_element: dict
        @param control_element: Implemented controls
        @type parent_component: str
        @param parent_component: UUID of parent component
        @rtype: dict
        @returns: New statement objects created
        """

        new_statements = []
        implemented_reqs = (
            control_element["implemented-requirements"]
            if "implemented-requirements" in control_element
            else []
        )
        for implemented_control in implemented_reqs:
            control_id = (
                implemented_control["control-id"]
                if "control-id" in implemented_control
                else "missing"
            )
            new_statement = Statement(
                sid=control_id,
                sid_class=catalog_key,
                pid=get_control_statement_part(control_id),
                source=catalog_key,
                uuid=control_element["uuid"] if "uuid" in control_element else uuid4(),
                body=implemented_control["description"]
                if "description" in implemented_control
                else "",
                statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name,
                remarks=implemented_control["remarks"]
                if "remarks" in implemented_control
                else "",
                status=implemented_control["status"]
                if "status" in implemented_control
                else None,
                producer_element=parent_component,
            )
            logger.info(f"New statement with UUID {new_statement.uuid} being created.")
            new_statements.append(new_statement)
        statements_created = Statement.objects.bulk_create(new_statements)
        return statements_created


def add_selected_components(system, import_record):
    """
    Add a component from the library or a compliance app to the project and its statements
    using the import record
    """

    # Get components from import record
    imported_components = Element.objects.filter(import_record=import_record)
    for imported_component in imported_components:
        for smt in Statement.objects.filter(
            producer_element_id=imported_component.id,
            statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name,
        ):
            smt.create_system_control_smt_from_component_prototype_smt(
                system.root_element.id
            )
    return imported_components


@login_required
def project_component_editor(
    request, system_id, element_id, catalog_key=None, control_id=None, statement_id=None
):
    """Display System's selected element detail view"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("view_system", system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        name = Element.objects.values_list("name", flat=True).filter(pk=element_id)
        parameters = project.root_task.module.app.catalog_metadata["parameters"]
        catalog_key = [p for p in parameters if p["id"] == "catalog_key"][0]["value"]

        statements = get_statements_by_control(
            system, element_id, catalog_key, statement_id
        )
        narrative = get_narrative(statements, statement_id)
        narrative["title"] = f'{narrative.get("label")}: {narrative.get("title")}'
        if narrative.get("next"):
            save = "Save & next"
            url = reverse(
                "system_element_control",
                args=[
                    system.id,
                    narrative.get("producer_element_id"),
                    narrative.get("control_id"),
                    catalog_key,
                    narrative.get("next"),
                ],
            )
        else:
            save = "Save"
            url = reverse(
                "system_element_control",
                args=[
                    system.id,
                    narrative.get("producer_element_id"),
                    narrative.get("control_id"),
                    catalog_key,
                    narrative.get("sid"),
                ],
            )
        narrative["save"] = {
            "text": save,
            "url": url,
        }

        cl_id = control_id if control_id else narrative.get("control_id")
        cat = get_catalog_data_by_control(catalog_key, cl_id)
        nav = project_navigation(request, project)

        # Return the system's element information
        context = {
            "title": name,
            "catalog": cat,
            "nav": nav,
            "project": project,
            "statements": statements,
            "system": system,
            "narrative": narrative,
        }
        return render(request, "components/component-editor.html", context)
    else:
        # User does not have permission to this system
        raise Http404


def edit_component_state(request, system_id, element_id):
    """
    Edit system component state
    """
    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("change_system", system):
        # Retrieve element
        # TODO: Make atomic transaction
        element = Element.objects.get(id=element_id)
        element.component_state = request.POST["state_change"]
        element.save()
        logger.info(
            event=f"change_system update_component_state {element} {element.component_state}",
            object={"object": "system", "id": system.id},
            user={"id": request.user.id, "username": request.user.username},
        )

        state_status = {
            "operational": "Implemented",
            "under-development": "Partially Implemented",
            "planned": "Planned",
        }
        control_status = (
            state_status.get(request.POST["state_change"]) or "Not Implemented"
        )
        system.set_component_control_status(element, control_status)
        logger.info(
            event=f"change_system batch_update_component_control_status {element} {control_status}",
            object={"object": "system", "id": system.id},
            user={"id": request.user.id, "username": request.user.username},
        )
    return redirect(reverse("system_element", args=[system_id, element_id]))


def edit_component_type(request, system_id, element_id):
    """
    Edit system component type
    """
    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("change_system", system):
        # Retrieve element
        element = Element.objects.get(id=element_id)
        element.component_type = request.POST["type_change"]
        element.save()
    return redirect(reverse("system_element", args=[system_id, element_id]))


@login_required
def system_element_remove(request, system_id, element_id):
    """Remove an element from a system and delete the controls it produces"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("change_system", system):

        # Retrieve element
        element = Element.objects.get(id=element_id)

        # Delete the control implementation statements associated with this component
        element.statements_produced.filter(
            consumer_element=system.root_element
        ).delete()

        # Log result
        logger.info(
            event="change_system remove_component",
            object={"object": "component", "id": element.id},
            user={"id": request.user.id, "username": request.user.username},
        )

        # Create message for user
        messages.add_message(
            request, messages.INFO, f"Component {element.name} was removed."
        )

    else:
        # User does not have permission
        # Log result
        logger.info(
            event="change_system remove_component permission_denied",
            object={"object": "component", "id": element_id},
            user={"id": request.user.id, "username": request.user.username},
        )

        # Create message for user
        messages.add_message(
            request, messages.INFO, "You don't have permission to edit the system."
        )

    response = redirect(reverse("components_selected", args=[system_id]))
    return response


@login_required
def new_element(request):
    """Form to create new system element (aka component)"""

    if request.method == "POST":
        form = ElementForm(request.POST)
        if form.is_valid():
            form.save()
            element = form.instance
            logger.info(
                event="new_element",
                object={"object": "element", "id": element.id, "name": element.name},
                user={"id": request.user.id, "username": request.user.username},
            )
            return redirect("component_library_component", element_id=element.id)
    else:
        form = ElementForm()

    return render(
        request,
        "components/element_form.html",
        {
            "form": form,
        },
    )


@login_required
def component_library_component(request, element_id, statement_id=None):
    """Display certified component's element detail view"""

    # Retrieve element
    element = Element.objects.get(id=element_id)
    if element.component_state:
        element.component_state = element.component_state.replace("_", " ")

    stmts = element.statements_produced.filter(
        statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name
    )
    catalog_key = stmts[0].sid_class
    catalog = Catalog.GetInstance(catalog_key=catalog_key)

    st = {}
    for s in stmts:
        active = False
        if statement_id and s.id == int(statement_id):
            active = True

        ctrl = catalog.get_control_by_id(s.sid)
        cid = catalog.get_control_property_by_name(ctrl, "sort-id")
        st[cid] = {
            "body": s.body,
            "inheritance": s.inheritance,
            "sid": s.id,
            "control_id": s.sid,
            "label": catalog.get_control_property_by_name(ctrl, "label"),
            "title": ctrl.get("title"),
            "producer_element_name": s.producer_element.name,
            "producer_element_id": s.producer_element.id,
            "status": s.status,
            "href": reverse(
                "component_library_component_details", args=[element.id, s.id]
            ),
            "active": active,
        }

    if st:
        statements = dict(sorted(st.items()))
        if not statement_id:
            k = list(statements.keys())[0]
            statements[k]["active"] = True
    else:
        statements = {}

    narrative = get_narrative(statements, statement_id)
    projects = Project.get_projects_with_read_priv(
        request.user, excludes={"contained_in_folders": None}
    )
    pids = []
    options = {}
    for p in projects:
        pids.append(p.system_id)
        options[p.system_id] = {"id": p.system_id, "title": p.title}

    root_ids = []
    systems = System.objects.filter(pk__in=pids)
    for s in systems:
        root_ids.append(s.root_element_id)
        cat_id = ElementControl.objects.filter(element_id=s.root_element_id).first()
        if not cat_id or cat_id.oscal_catalog_key != catalog_key:
            options.pop(s.id)
    existing_list = (
        Statement.objects.filter(consumer_element_id__in=root_ids)
        .filter(producer_element_id=element_id)
        .values_list("consumer_element_id", flat=True)
        .distinct()
        .order_by("consumer_element_id")
    )

    cid = narrative.get("control_id")
    cat = get_catalog_data_by_control(catalog_key, cid)

    context = {
        "element": element,
        "catalog": cat,
        "catalog_key": catalog_key,
        "narrative": narrative,
        "statements": statements,
        "options": options,
        "existing_list": existing_list,
    }
    return render(request, "components/library/component_details.html", context)


@login_required
def api_controls_select(request):
    """Return list of controls in json for select2 options from all control catalogs"""

    cl_id = request.GET.get("q", None)
    oscal_ctl_id = oscalize_control_id(cl_id)
    catalogs_containing_cl_id = CatalogData.objects.filter(
        Q(
            catalog_json__catalog__groups__contains=[
                {"controls": [{"id": oscal_ctl_id}]}
            ]
        )
        | Q(
            catalog_json__catalog__groups__contains=[
                {"controls": [{"controls": [{"id": oscal_ctl_id}]}]}
            ]
        )
    )
    cxs = []
    for catalog in catalogs_containing_cl_id:
        catalog_key_display = catalog.catalog_key.replace("_", " ")
        # TODO: get control title effectively from CatalogData
        # title = "catalog.ctl_title"
        cxs.append(
            {
                "id": oscal_ctl_id,
                "catalog_key_display": catalog_key_display,
                "display_text": f"{oscal_ctl_id} - {catalog_key_display} - {cl_id}",
            }
        )
    status = "success"
    message = "Sending list."
    return JsonResponse(
        {"status": status, "message": message, "data": {"controls": cxs}}
    )


@login_required
def component_library_component_copy(request, element_id):
    """Copy a component"""

    # Retrieve element
    element = Element.objects.get(id=element_id)
    count = Element.objects.filter(uuid=element.uuid).count()

    if count > 0:
        e_copy = element.copy(name=element.name + " copy (" + str(count + 1) + ")")
    else:
        e_copy = element.copy()

    # Create message to display to user
    messages.add_message(
        request, messages.INFO, f"Component {element.name} was copied to {e_copy.name}"
    )

    # # Redirect to the new page for the component
    return HttpResponseRedirect("/controls/components/{}".format(e_copy.id))


@login_required
def import_component(request):
    """Import a Component in JSON"""

    import_name = request.POST.get("import_name", "")
    oscal_component_json = request.POST.get("json_content", "")
    ComponentImporter().import_components_as_json(
        import_name, oscal_component_json, request
    )
    return component_library(request)


def raise_404_if_not_permitted_to_statement(
    request, statement, system_permission="view_system"
):
    """Raises a 404 if the user doesn't have the statement system permission"""
    while True:
        if request.user.is_superuser or request.user.is_staff:
            # Allow access if the user has superuser or staff member permissions respectively
            break
        else:
            # Allow access if the user has the statement system permission
            try:
                system = System.objects.get(root_element=statement.consumer_element)
                if request.user.has_perm(system_permission, system):
                    break
            except System.DoesNotExist:
                logger.info(
                    event="update_smt_permission permission_denied",
                    object={"object": "statement", "id": statement.id},
                    user={"id": request.user.id, "username": request.user.username},
                )
                raise Http404


@login_required
def statement_history(request, smt_id=None):
    """Returns the history for the specified statement"""

    # Get statement if exists else 404
    smt = get_object_or_404(Statement, id=smt_id)

    # Check permission block
    permission = False
    if request.user.is_superuser or request.user.is_staff:
        # Grant permission to superusers and staff users
        permission = True
    elif System.objects.filter(root_element=smt.consumer_element).exists():
        # Grant permission to user with edit access on system
        system = System.objects.get(root_element=smt.consumer_element)
        permission = True if request.user.has_perm("view_system", system) else False
    # 404 if user does not have permission
    if not permission:
        raise Http404

    # Check permission
    raise_404_if_not_permitted_to_statement(request, smt)

    return render(
        request, "controls/statement_history.html", {"statement": smt.history.all()}
    )


@login_required
def restore_to_history(request, smt_id, history_id):
    """Restores the specified statement version"""

    # Get statement if exists else 404
    smt = get_object_or_404(Statement, id=smt_id)

    # Check permission
    raise_404_if_not_permitted_to_statement(request, smt, "change_system")

    full_smt_history = None
    for query_key in request.POST:
        if "restore" in query_key:
            change_reason = request.POST.get(query_key, "")
        else:
            change_reason = None

    try:
        # Save historical statement as a new instance
        historical_smt = smt.history.get(history_id=history_id)
        historical_smt.instance.save()

        # Update the reason for the new statement record
        recent_smt = smt.history.first()
        update_change_reason(recent_smt.instance, change_reason)

        logger.info(f"Change reason: {change_reason}")

        logger.info(
            f"Restoring the current statement with an id of {smt_id} to version with a \
                history id of {history_id}"
        )
        messages.add_message(
            request,
            messages.INFO,
            f"The statement was restored to statement version {history_id}",
        )

        # Diff between most recent and the historical record
        full_smt_history = smt.history.all()
        recent_record = full_smt_history.filter(
            history_id=recent_smt.history_id
        ).first()
        historical_record = full_smt_history.filter(
            history_id=historical_smt.history_id
        ).first()

        delta = historical_record.diff_against(recent_record)
        for change in delta.changes:
            logger.info(
                "{} changed from {} to {}".format(change.field, change.old, change.new)
            )
    except ObjectDoesNotExist as ex:
        messages.add_message(
            request,
            messages.ERROR,
            f"{ex} No statement found. Please confirm your entry is correct and try again.",
        )

    context = {
        "history_id": history_id,
        "smt_id": smt_id,
        "statement": full_smt_history,
    }

    return render(request, "controls/statement_history.html", context)


@login_required
def system_element_download_oscal_json(request, system_id, element_id):

    if system_id is not None and system_id != "":
        # Retrieve identified System
        system = System.objects.get(id=system_id)
        # Retrieve related selected controls if user has permission on system
        if request.user.has_perm("view_system", system):
            # Retrieve element
            element = Element.objects.get(id=element_id)

            # Retrieve impl_smts produced by element and consumed by system
            # Get the impl_smts contributed by this component to system
            impl_smts = element.statements_produced.filter(
                consumer_element=system.root_element
            )
    else:
        # Comes from Component Library, no system

        # Retrieve element
        element = Element.objects.get(id=element_id)
        # Get the impl_smts contributed by this component to system
        impl_smts = Statement.objects.filter(producer_element=element)
    print(f"component library statements {len(impl_smts)}")
    response = HttpResponse(content_type="application/json")
    filename = str(PurePath(slugify(element.name)).with_suffix(".json"))
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    body = OSCALComponentSerializer(element, impl_smts).as_json()
    response.write(body)

    return response


@login_required
def OSCAL_ssp_export(*args, **kwargs):
    """
    Exporting a system security plan in OSCAL json version 1.0.0
    """
    system_id = kwargs.get("system_id", 1)
    # Retrieve identified System
    system = System.objects.get(id=system_id)
    impl_smts = Statement.objects.filter(
        consumer_element=system.root_element,
        statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION.name,
    )
    oscal_string = OSCALSystemSecurityPlanSerializer(system, impl_smts).as_json()
    # File name construction and JSON response
    filename = "{}_OSCAL_{}.json".format(
        system.root_element.name.replace(" ", "_"),
        datetime.now().strftime("%Y-%m-%d-%H-%M"),
    )
    resp = HttpResponse(oscal_string, content_type="application/json")
    resp["content-disposition"] = "attachment; filename=%s" % quote(filename)
    return resp


@login_required
def controls_selected_export_xacta_xslx(request, system_id):
    """Export System's selected controls compatible with Xacta 360"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("view_system", system):
        # Retrieve primary system Project
        controls = system.root_element.controls.all()

        # Retrieve any related Implementation Statements
        impl_smts = system.root_element.statements_consumed.all()
        impl_smts_by_sid = {}
        for smt in impl_smts:
            if smt.sid in impl_smts_by_sid:
                impl_smts_by_sid[smt.sid].append(smt)
            else:
                impl_smts_by_sid[smt.sid] = [smt]

        for control in controls:
            if control.oscal_ctl_id in impl_smts_by_sid:
                setattr(control, "impl_smts", impl_smts_by_sid[control.oscal_ctl_id])
            else:
                setattr(control, "impl_smts", None)

        from tempfile import NamedTemporaryFile

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Controls_Implementation"

        # Add in field name row
        # Paragraph/ReqID
        c = ws.cell(row=1, column=1, value="Paragraph/ReqID")
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            left=Side(border_style="thin", color="444444"),
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Stated Requirement (Control statement/Requirement)
        c = ws.cell(row=1, column=2, value="Title")
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions["B"].width = 30
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Private Implementation
        c = ws.cell(row=1, column=3, value="Private Implementation")
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions["C"].width = 80
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Public Implementation
        c = ws.cell(row=1, column=4, value="Public Implementation")
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        ws.column_dimensions["D"].width = 80
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Notes
        c = ws.cell(row=1, column=5, value="Notes")
        ws.column_dimensions["E"].width = 60
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Status ["Implemented", "Planned"]
        c = ws.cell(row=1, column=6, value="Status")
        ws.column_dimensions["F"].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Expected Completion (expected implementation)
        c = ws.cell(row=1, column=7, value="Expected Completion")
        ws.column_dimensions["G"].width = 20
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Class ["Management", "Operational", "Technical",
        c = ws.cell(row=1, column=8, value="Class")
        ws.column_dimensions["H"].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Priority ["p0", "P1", "P2", "P3"]
        c = ws.cell(row=1, column=9, value="Priority")
        ws.column_dimensions["I"].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Responsible Entities
        c = ws.cell(row=1, column=10, value="Responsible Entities")
        ws.column_dimensions["J"].width = 20
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Control Owner(s)
        c = ws.cell(row=1, column=11, value="Control Owner(s)")
        ws.column_dimensions["K"].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Type ["System-Specific", "Hybrid", "Inherited", "Common", "blank"]
        c = ws.cell(row=1, column=12, value="Type")
        ws.column_dimensions["L"].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Inherited From
        c = ws.cell(row=1, column=13, value="Inherited From")
        ws.column_dimensions["M"].width = 20
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Provide As ["Do Not Share", "blank"]
        c = ws.cell(row=1, column=14, value="Provide As")
        ws.column_dimensions["N"].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Evaluation Status ["Evaluated", "Expired", "Not Evaluated", "Unknown", "blank"]
        c = ws.cell(row=1, column=15, value="Evaluation Status")
        ws.column_dimensions["O"].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # Control Origination
        c = ws.cell(row=1, column=16, value="Control Origination")
        ws.column_dimensions["P"].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        # History
        c = ws.cell(row=1, column=17, value="History")
        ws.column_dimensions["Q"].width = 15
        c.fill = PatternFill("solid", fgColor="5599FE")
        c.font = Font(color="FFFFFF", bold=True)
        c.border = Border(
            right=Side(border_style="thin", color="444444"),
            bottom=Side(border_style="thin", color="444444"),
            outline=Side(border_style="thin", color="444444"),
        )

        for row in range(2, len(controls) + 1):
            control = controls[row - 2]

            # Paragraph/ReqID
            c = ws.cell(
                row=row,
                column=1,
                value=control.get_flattened_oscal_control_as_dict()[
                    "id_display"
                ].upper(),
            )
            c.fill = PatternFill("solid", fgColor="FFFF99")
            c.alignment = Alignment(vertical="top", wrapText=True)
            c.border = Border(
                left=Side(border_style="thin", color="444444"),
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Title
            c = ws.cell(
                row=row,
                column=2,
                value=control.get_flattened_oscal_control_as_dict()["title"],
            )
            c.fill = PatternFill("solid", fgColor="FFFF99")
            c.alignment = Alignment(vertical="top", wrapText=True)
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Private Implementation
            smt_combined = ""
            if control.impl_smts:
                for smt in control.impl_smts:
                    smt_combined += smt.body
            c = ws.cell(row=row, column=3, value=smt_combined)
            c.alignment = Alignment(vertical="top", wrapText=True)
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Public Implementation
            c.alignment = Alignment(vertical="top", wrapText=True)
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Notes
            c = ws.cell(row=1, column=5, value="Notes")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Status ["Implemented", "Planned"]
            c = ws.cell(row=1, column=6, value="Status")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Expected Completion (expected implementation)
            c = ws.cell(row=1, column=7, value="Expected Completion")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Class ["Management", "Operational", "Technical",
            c = ws.cell(row=1, column=8, value="Class")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Priority ["p0", "P1", "P2", "P3"]
            c = ws.cell(row=1, column=9, value="Priority")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Responsible Entities
            c = ws.cell(row=1, column=10, value="Responsible Entities")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Control Owner(s)
            c = ws.cell(row=1, column=11, value="Control Owner(s)")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Type ["System-Specific", "Hybrid", "Inherited", "Common", "blank"]
            c = ws.cell(row=1, column=12, value="Type")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Inherited From
            c = ws.cell(row=1, column=13, value="Inherited From")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Provide As ["Do Not Share", "blank"]
            c = ws.cell(row=1, column=14, value="Provide As")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Evaluation Status ["Evaluated", "Expired", "Not Evaluated", "Unknown", "blank"]
            c = ws.cell(row=1, column=15, value="Evaluation Status")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # Control Origination
            c = ws.cell(row=1, column=16, value="Control Origination")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

            # History
            c = ws.cell(row=1, column=17, value="History")
            c.border = Border(
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
            blob = stream

        mime_type = "application/octet-stream"
        filename = "{}_control_implementations-{}.xlsx".format(
            system.root_element.name.replace(" ", "_"),
            datetime.now().strftime("%Y-%m-%d-%H-%M"),
        )

        resp = HttpResponse(blob, mime_type)
        resp["Content-Disposition"] = "inline; filename=" + filename
        return resp
    else:
        # User does not have permission to this system
        raise Http404


@login_required
def project_control_editor(request, system_id, catalog_key, cl_id, statement_id=None):
    """System Control detail view"""

    catalog_key, system = get_editor_system(catalog_key, system_id)

    # Retrieve related statements if user has permission on system
    if request.user.has_perm("view_system", system):
        project = system.projects.first()
        catalog = Catalog.GetInstance(catalog_key=catalog_key)
        cg_flat = catalog.get_flattened_controls_all_as_dict()
        # If control id does not exist in catalog
        if cl_id.lower() not in cg_flat:
            return render(
                request, "controls/detail.html", {"catalog": catalog, "control": {}}
            )

        statements = get_statements_by_component(
            system, cl_id, catalog_key, statement_id
        )
        if statements:
            narrative = get_narrative(statements, statement_id)
            narrative["title"] = narrative.get("producer_element_name")
            if narrative.get("next"):
                save = "Save & next"
                url = reverse(
                    "control_editor_statement",
                    args=[system.id, catalog_key, cl_id, narrative.get("next")],
                )
            else:
                save = "Save"
                url = reverse(
                    "control_editor_statement",
                    args=[system.id, catalog_key, cl_id, narrative.get("sid")],
                )
            narrative["save"] = {
                "text": save,
                "url": url,
            }
        else:
            narrative = {}

        cat = get_catalog_data_by_control(catalog_key, cl_id)
        nav = project_navigation(request, project)

        context = {
            "catalog": cat,
            "nav": nav,
            "project": project,
            "statements": statements,
            "system": system,
            "narrative": narrative,
        }
        return render(request, "controls/editor.html", context)
    else:
        # User does not have permission to this system
        raise Http404


def get_catalog_data_by_control(catalog_key, control_id):
    """
    Return all of the data associated with a given control for a given catalog
    """
    cat = Catalog.GetInstance(catalog_key=catalog_key)
    ctrl = cat.get_control_by_id(control_id)
    catalog_data = {
        "control": cat.get_flattened_control_as_dict(ctrl),
        "description": cat.get_control_prose_as_markdown(ctrl, "statement"),
        "guidance": cat.get_control_prose_as_markdown(ctrl, "guidance"),
        "implementation": cat.get_control_prose_as_markdown(ctrl, "implementation"),
        "catalog_display": cat.catalog_key_display,
        "catalog_key": catalog_key,
    }
    return catalog_data


def get_statements_by_component(system, control_id, catalog_key, statement_id):
    """
    Given a system element, a control ID and a catalog key, return the associated
    statements.
    """

    stmts = Statement.objects.filter(
        sid=control_id,
        consumer_element=system.root_element,
        sid_class=catalog_key,
        statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION.name,
    ).order_by("pid")

    st = {}
    for s in stmts:
        active = False
        if statement_id and s.id == int(statement_id):
            active = True

        st[s.producer_element.name] = {
            "body": s.body,
            "inheritance": s.inheritance,
            "label": s.producer_element.name,
            "sid": s.id,
            "producer_element_name": s.producer_element.name,
            "producer_element_id": s.producer_element.id,
            "status": s.status,
            "href": reverse(
                "control_editor_statement",
                args=[system.id, catalog_key, control_id, s.id],
            ),
            "active": active,
        }

    if st:
        statements = dict(sorted(st.items()))
        if not statement_id:
            k = list(statements.keys())[0]
            statements[k]["active"] = True
    else:
        statements = {}

    return statements


def get_statements_by_control(system, component_id, catalog_key, statement_id):
    """
    Given a system element, a control ID and a catalog key, return the associated
    statements.
    """

    stmts = Statement.objects.filter(
        producer_element_id=component_id,
        consumer_element=system.root_element,
        sid_class=catalog_key,
        statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION.name,
    ).order_by("pid")

    catalog = Catalog.GetInstance(catalog_key=catalog_key)

    st = {}
    for s in stmts:
        active = False
        if statement_id and s.id == int(statement_id):
            active = True

        ctrl = catalog.get_control_by_id(s.sid)
        cid = catalog.get_control_property_by_name(ctrl, "sort-id")
        st[cid] = {
            "body": s.body,
            "inheritance": s.inheritance,
            "sid": s.id,
            "control_id": s.sid,
            "label": catalog.get_control_property_by_name(ctrl, "label"),
            "title": ctrl.get("title"),
            "producer_element_name": s.producer_element.name,
            "producer_element_id": s.producer_element.id,
            "status": s.status,
            "href": reverse(
                "system_element_control",
                args=[system.id, component_id, catalog_key, s.sid, s.id],
            ),
            "active": active,
        }

    if st:
        statements = dict(sorted(st.items()))
        if not statement_id:
            k = list(statements.keys())[0]
            statements[k]["active"] = True
    else:
        statements = {}

    return statements


def get_narrative(statements, statement_id):
    """
    Given a statement_id, or None if one wasn't provided, determine which item
    in the statements dict to display and determine whether this is the last
    item in the dict.

    :param dict statements: A dictionary containing the project specific
    Control statements for a given Control keyed by the Component name.
    :param int statement_id: The Statement ID for the narrative to display
    and edit.
    :return dict statements: Return a dict containing the values associate with
    the Control narrative that is to be displayed/edited or return None.
    """
    statements = list(statements.values())

    # add a "next" key to statements linking them to the "sid" of the
    # following statement
    last_idx = len(statements) - 1
    for idx, statement in enumerate(statements):
        if idx < last_idx:
            statement["next"] = statements[idx + 1]["sid"]

    # if called with a statement_id, we're only interested in a subset
    if statement_id:
        statements = [s for s in statements if s["sid"] == int(statement_id)]

    # return the first statement, or None if there are no matching statements
    if statements:
        return statements[0]
    else:
        return None


def get_editor_system(catalog_key, system_id):
    """
    Retrieves oscalized control id and catalog key. Also system object from system id.
    """

    catalog_key = oscalize_catalog_key(catalog_key)

    # Retrieve identified System
    system = System.objects.get(id=system_id)

    return catalog_key, system


@login_required
def editor_compare(request, system_id, catalog_key, cl_id):
    """System Control detail view"""

    catalog_key, system = get_editor_system(catalog_key, system_id)
    # oscalize key
    cl_id = oscalize_control_id(cl_id)
    # Retrieve related statements if owner has permission on system
    if request.user.has_perm("view_system", system):
        project, catalog, cg_flat, impl_smts = get_catalog_data_by_control(
            request, system, catalog_key, cl_id
        )
        context = {
            "system": system,
            "project": project,
            "catalog": catalog,
            "control": cg_flat[cl_id.lower()],
            "impl_smts": impl_smts,
        }
        return render(request, "controls/compare.html", context)
    else:
        # User does not have permission to this system
        raise Http404


@login_required
def get_editor_data(request, system, catalog_key, cl_id):
    if request.user.has_perm("view_system", system):
        project = system.projects.first()
        parameter_values = project.get_parameter_values(catalog_key)
        catalog = Catalog(catalog_key, parameter_values=parameter_values)
        cg_flat = catalog.get_flattened_controls_all_as_dict()
        if cl_id.lower() not in cg_flat:
            return render(
                request, "controls/detail.html", {"catalog": catalog, "control": {}}
            )

        impl_smts = Statement.objects.filter(
            sid=cl_id,
            consumer_element=system.root_element,
            sid_class=catalog_key,
            statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION.name,
        ).order_by("pid")
        # Retrieve Legacy Implementation Statements
        impl_smts_legacy = Statement.objects.filter(
            sid=cl_id,
            consumer_element=system.root_element,
            sid_class=catalog_key,
            statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_LEGACY.name,
        )
        return project, catalog, cg_flat, impl_smts, impl_smts_legacy


# @task_view
@login_required
def save_smt(request):
    """Save a statement"""

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    else:
        # Track if we are creating a new statement
        new_statement = False
        form_dict = dict(request.POST)
        form_values = {}
        for key in form_dict.keys():
            form_values[key] = form_dict[key][0]
        smt_id = form_values["smt_id"]
        # Updating or saving a new statement?
        if len(smt_id) > 0:
            # Look up existing Statement object
            statement = Statement.objects.get(pk=smt_id)
            # Check if statement has the same sid class as the statement object
            if statement.sid_class == form_values["sid_class"]:
                # Check user permissions
                system = statement.consumer_element
                if not request.user.has_perm("change_system", system):
                    # User does not have write permissions
                    # Log permission to save answer denied
                    logger.info(
                        event="save_smt permission_denied",
                        object={"object": "statement", "id": statement.id},
                        user={"id": request.user.id, "username": request.user.username},
                    )
                    return HttpResponseForbidden(
                        f"Permission denied. {request.user.username} does not have change \
                            privileges to system and/or project."
                    )

                if statement is None:
                    # Statement from received has an id no longer in the database.
                    # Report error. Alternatively, in future save as new Statement object
                    statement_status = "error"
                    statement_msg = (
                        "The id for this statement is no longer valid in the database."
                    )
                    return JsonResponse(
                        {"status": statement_status, "message": statement_msg}
                    )
                # Update existing Statement object with received info
                statement.pid = form_values["pid"]
                statement.body = form_values["body"]
                statement.remarks = form_values["remarks"]
                if "status" in form_values:
                    statement.status = form_values["status"]
                else:
                    statement.status = None
                if "inheritance_name" in form_values:
                    statement.inheritance_id = form_values["inheritance_name"]
            else:
                new_statement = True
        else:
            new_statement_type_enum = StatementTypeEnum[
                form_values["statement_type"].upper()
            ]
            # Create new Statement object
            new_sid_class = form_values["sid_class"].replace(
                " ", "_"
            )  # convert displayed catalog name to catalog_key
            statement = Statement(
                sid=oscalize_control_id(form_values["sid"]),
                sid_class=new_sid_class,
                source=new_sid_class,
                body=form_values["body"],
                pid=form_values["pid"],
                statement_type=new_statement_type_enum.name,
                status=form_values["status"],
                remarks=form_values["remarks"],
                consumer_element_id=form_values["consumer_element_id"],
            )
            new_statement = True

        # Updating or saving a new producer_element?
        try:
            # Does the name match and existing element? (Element names are unique.)
            # TODO: Sanitize data entered in form?
            producer_element, created = Element.objects.get_or_create(
                name=form_values["producer_element_name"]
            )
            if created:
                producer_element.element_type = "system_element"
                producer_element.save()
            producer_element_status = "ok"
            producer_element_msg = "Producer Element saved."
        except Exception as e:
            producer_element_status = "error"
            producer_element_msg = (
                "Producer Element save failed. Error reported {}".format(e)
            )
            return JsonResponse(
                {"status": producer_element_status, "message": producer_element_msg}
            )

        # Associate Statement and Producer Element if creating new statement
        if new_statement:
            try:
                statement.producer_element = producer_element
                statement.save()
                statement_element_status = "ok"
                statement_element_msg = (
                    "Control statement associated with Producer Element"
                )
                messages.add_message(
                    request,
                    messages.INFO,
                    f"{statement_element_msg} {producer_element.id} was added.",
                )
            except Exception as e:
                statement_element_status = "error"
                statement_element_msg = (
                    "Failed to associate statement with Producer Element {}".format(e)
                )
                return JsonResponse(
                    {
                        "status": statement_element_status,
                        "message": statement_element_msg
                        + " "
                        + producer_element_msg
                        + " "
                        + statement_element_msg,
                    }
                )
        # Create new Prototype Statement object on new statement creation (not statement edit)
        if new_statement:
            try:
                statement.status = None
            except Exception as e:
                statement_status = "error"
                statement_msg = f"Statement save failed while saving statement prototype. \
                    Error reported {e}"
                return JsonResponse(
                    {"status": statement_status, "message": statement_msg}
                )

        statement_del_msg = ""
        if (
            "form_source" in form_values
            and form_values["form_source"] == "component_library"
        ):
            # Form received from component library
            from django.core import serializers

            serialized_obj = serializers.serialize(
                "json",
                [
                    statement,
                ],
            )
            # Delete statement
            Statement.objects.filter(pk=statement.id).delete()
            statement.delete()
            statement_del_msg = "Orphaned Control_Implementation Statement deleted."
        else:
            # Associate Statement and System's root_element
            system_id = form_values["system_id"]
            if new_statement and system_id is not None:
                try:
                    statement.consumer_element = System.objects.get(
                        pk=form_values["system_id"]
                    ).root_element
                    # statement.save()
                    statement_msg = "Statement associated with System/Consumer Element."
                except Exception as e:
                    statement_consumer_status = "error"
                    statement_consumer_msg = f"Failed to associate statement with System/\
                        Consumer Element {e}"
                    return JsonResponse(
                        {
                            "status": statement_consumer_status,
                            "message": statement_msg
                            + " "
                            + producer_element_msg
                            + " "
                            + statement_consumer_msg,
                        }
                    )

            # Serialize saved data object(s) to send back to update web page
            # The submitted form needs to be updated with the object primary keys (ids)
            # in order that future saves will be treated as updates.
            from django.core import serializers

            serialized_obj = serializers.serialize(
                "json",
                [
                    statement,
                ],
            )

    # Save Statement object
    try:
        if not new_statement:
            statement.save()
        statement_msg = "Statement saved."
        messages.add_message(request, messages.INFO, f"Statement {smt_id}  was saved.")
    except Exception as e:
        statement_status = "error"
        statement_msg = "Statement save failed. Error reported {}".format(e)
        return JsonResponse({"status": statement_status, "message": statement_msg})
    # Return successful save result to web page's Ajax request
    return JsonResponse(
        {
            "status": "success",
            "message": statement_msg
            + " "
            + producer_element_msg
            + " "
            + statement_del_msg,
            "statement": serialized_obj,
        }
    )


@login_required
def update_smt_prototype(request):
    """Update a certified statement"""

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    else:
        # Get statement if exists else 404
        form_dict = dict(request.POST)
        form_values = {}
        for key in form_dict.keys():
            form_values[key] = form_dict[key][0]
        smt_id = form_values["smt_id"]
        statement = get_object_or_404(Statement, pk=smt_id)

        # Check permission
        raise_404_if_not_permitted_to_statement(request, statement)

        if statement is None:
            statement_msg = (
                "The id for this statement is no longer valid in the database."
            )
            return JsonResponse({"status": "error", "message": statement_msg})

        # needs self.body == self.prototype.body

        try:
            proto_statement = Statement.objects.get(pk=statement.prototype_id)
            proto_statement.prototype.body = statement.body
            proto_statement.prototype.save()
            statement_msg = f"Update to statement prototype {proto_statement.prototype_id} \
                succeeded."
        except Exception as e:
            statement_msg = (
                "Update to statement prototype failed. Error reported {}".format(e)
            )
            return JsonResponse({"status": "error", "message": statement_msg})

        return JsonResponse(
            {
                "status": "success",
                "message": statement_msg,
                "data": {"smt_body": statement.body},
            }
        )


@login_required
def delete_smt(request):
    """Delete a statement"""

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    else:
        form_dict = dict(request.POST)
        form_values = {}
        for key in form_dict.keys():
            form_values[key] = form_dict[key][0]
        smt_id = form_values["smt_id"]
        # Delete statement?
        statement = Statement.objects.get(pk=smt_id)

        # Check user permissions
        system = statement.consumer_element
        if not request.user.has_perm("change_system", system):
            # User does not have write permissions
            # Log permission to save answer denied
            logger.info(
                event="delete_smt permission_denied",
                object={"object": "statement", "id": statement.id},
                user={"id": request.user.id, "username": request.user.username},
            )
            return HttpResponseForbidden(
                f"Permission denied. {request.user.username} does not have change privileges \
                    to system and/or project."
            )

        if statement is None:
            # Statement from received has an id no longer in the database.
            # Report error. Alternatively, in future save as new Statement object
            statement_msg = (
                "The id for this statement is no longer valid in the database."
            )
            return JsonResponse({"status": "error", "message": statement_msg})
        # Delete Statement object
        try:
            statement.delete()
            statement_msg = "Statement deleted."
        except Exception as e:
            statement_msg = f"Statement delete failed. Error reported {e}"
            return JsonResponse({"status": "error", "message": statement_msg})

        return JsonResponse({"status": "success", "message": statement_msg})


# Components


@login_required
def add_system_component(request):
    """Add an existing element and its statements to a system"""

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    form_dict = dict(request.POST)
    form_values = {}
    for key in form_dict.keys():
        form_values[key] = form_dict[key][0]
    print(f"form_values: {form_values}")
    system_id = form_values["system_id"]
    producer_element_id = form_values["producer_element_id"]
    redirect_url = request.META.get("HTTP_REFERER")
    # Does user have permission to add element?
    # Check user permissions
    system = System.objects.get(pk=system_id)
    if not request.user.has_perm("change_system", system):
        # User does not have write permissions
        # Log permission to save answer denied
        logger.info(
            event="change_system permission_denied",
            object={
                "object": "element",
                "producer_element_name": form_values["producer_element_name"],
            },
            user={"id": request.user.id, "username": request.user.username},
        )
        return HttpResponseForbidden(
            f"Permission denied. {request.user.username} does not have change privileges to \
                system and/or project."
        )

    # Get system's existing components selected
    elements_selected = system.producer_elements
    elements_selected_ids = [e.id for e in elements_selected]

    # Add element to system's selected components
    # Look up the element rto add
    producer_element = Element.objects.get(pk=form_values["producer_element_id"])

    # TODO: various use cases
    # - component previously added but element has statements not yet added to system
    #   this issue may be best addressed elsewhere.

    # Component already added to system. Do not add the component (element) to the system again.
    if producer_element.id in elements_selected_ids:
        messages.add_message(
            request,
            messages.ERROR,
            f"Component {producer_element.name} can't be added because it is already included \
                in your system components.",
        )
        # Redirect to selected element page
        return HttpResponseRedirect(redirect_url)

    smts = Statement.objects.filter(
        producer_element_id=producer_element.id,
        statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION_PROTOTYPE.name,
    )

    # Component does not have any statements of type control_implementation_prototype to
    # add to system. So we cannot add the component (element) to the system.
    if len(smts) == 0:
        messages.add_message(
            request,
            messages.ERROR,
            f"{producer_element.name} can't be added because it does not have any control \
                implementation statements to add.",
        )
        # Redirect to selected element page
        return HttpResponseRedirect(redirect_url)

    # Loop through all element's prototype statements and add to control implementation statements.
    # System's selected controls will filter what controls and control statements to display.
    for smt in smts:
        smt.create_system_control_smt_from_component_prototype_smt(
            system.root_element.id
        )

    # Make sure some controls were added to the system. Report error otherwise.
    smts_added = Statement.objects.filter(
        producer_element_id=producer_element.id,
        consumer_element_id=system.root_element.id,
        statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION.name,
    )

    smts_added_count = len(smts_added)
    if smts_added_count > 0:
        messages.add_message(
            request,
            messages.INFO,
            f"{producer_element.name} and its {smts_added_count} control implementation \
                statements were added.",
        )
    else:
        messages.add_message(
            request,
            messages.WARNING,
            f"{producer_element.name} wasn't added to the system because it has no controls.",
        )
    projects = Project.get_projects_with_read_priv(
        request.user, excludes={"contained_in_folders": None}
    )
    pids = []
    options = {}
    for p in projects:
        pids.append(p.system_id)
        options[p.system_id] = {"id": p.system_id, "title": p.title}

    root_ids = []
    systems = System.objects.filter(pk__in=pids)
    for s in systems:
        root_ids.append(s.root_element_id)
        options[s.id]["pid"] = s.root_element_id

    existing_list = (
        Statement.objects.filter(consumer_element_id__in=root_ids)
        .filter(producer_element_id=producer_element_id)
        .values_list("consumer_element_id", flat=True)
        .distinct()
        .order_by("consumer_element_id")
    )
    context = {"options": options, "existing_list": existing_list}
    request.context = context
    # Redirect to selected element page
    # return HttpResponseRedirect(redirect_url)
    return HttpResponseRedirect(redirect_url)


@login_required
def search_system_component(request):
    """Add an existing element and its statements to a system"""

    if request.method != "GET":
        return HttpResponseNotAllowed(["GET"])

    form_dict = dict(request.GET)
    form_values = {}
    for key in form_dict.keys():
        form_values[key] = form_dict[key][0]
    # Form values from ajax data

    if "system_id" in form_values.keys():
        system_id = form_values["system_id"]

        # Does user have permission to add element?
        # Check user permissions
        system = System.objects.get(pk=system_id)

        if not request.user.has_perm("change_system", system):
            # User does not have write permissions
            # Log permission to save answer denied
            logger.info(
                event="change_system permission_denied",
                object={
                    "object": "element",
                    "producer_element_name": form_values["producer_element_name"],
                },
                user={"id": request.user.id, "username": request.user.username},
            )
            return HttpResponseForbidden(
                f"Permission denied. {request.user.username} does not have change privileges \
                    to system and/or project."
            )

        selected_controls = system.root_element.controls.all()
        selected_controls_ids = set()
        for sc in selected_controls:
            selected_controls_ids.add(
                "{} {}".format(sc.oscal_ctl_id, sc.oscal_catalog_key)
            )

        # Add element
        # Look up the element

        text = form_values["text"]

        # The final elements that are returned to the new dropdown created...
        producer_system_elements = Element.objects.filter(
            element_type="system_element"
        ).filter(name__contains=text)

        producer_elements = [
            {"id": str(ele.id), "name": ele.name} for ele in producer_system_elements
        ]

        results = {"producer_element_statement_values": producer_elements}
        data = json.dumps(results)
        mimetype = "application/json"
        return HttpResponse(data, mimetype)


class RelatedComponentStatements(View):
    """
    Returns the component statements that are produced(related) to one control
    implementation prototype
    """

    template_name = "controls/editor.html"

    def get(self, request):
        """Add an existing element and its statements to a system"""

        if request.method != "GET":
            return HttpResponseNotAllowed(["GET"])

        logger.info(f"Related controls GET: {dict(request.GET)}")
        form_dict = dict(request.GET)
        form_values = {}
        for key in form_dict.keys():
            form_values[key] = form_dict[key][0]
        # Form values from ajax data

        if "system_id" in form_values.keys():
            system_id = form_values["system_id"]

            # Does user have permission to add element?
            # Check user permissions
            system = System.objects.get(pk=system_id)

            if not request.user.has_perm("change_system", system):
                # User does not have write permissions
                # Log permission to save answer denied
                logger.info(
                    event="change_system permission_denied",
                    object={
                        "object": "element",
                        "producer_element_name": form_values["producer_element_name"],
                    },
                    user={"id": request.user.id, "username": request.user.username},
                )
                return HttpResponseForbidden(
                    f"Permission denied. {request.user.username} does not have change \
                        privileges to system and/or project."
                )

            selected_controls = system.root_element.controls.all()
            selected_controls_ids = set()
            for sc in selected_controls:
                selected_controls_ids.add(
                    "{} {}".format(sc.oscal_ctl_id, sc.oscal_catalog_key)
                )

            # Add element
            # Look up the element
            producer_element_id = form_values["producer_element_form_id"]
            producer_element = Element.objects.get(pk=producer_element_id)

            # The final elements that are returned to the new dropdown created...
            producer_smt_imps = producer_element.statements("control_implementation")

            producer_smt_imps_vals = [
                {
                    "smt_id": str(smt.id),
                    "smt_sid": smt.sid,
                    "smt_sid_class": smt.sid_class,
                    "producer_element_name": producer_element.name,
                    "smt_body": str(smt.body),
                    "smt_pid": str(smt.pid),
                    "smt_status": str(smt.status),
                }
                for smt in producer_smt_imps
            ]

            results = {
                "producer_element_statement_values": producer_smt_imps_vals,
                "selected_component": producer_element.name,
            }
            data = json.dumps(results)
            mimetype = "application/json"
            if data:
                return HttpResponse(data, mimetype)
            else:
                return JsonResponse(
                    status=400, data={"status": "error", "message": f"No DATA!: {data}"}
                )
        else:
            return JsonResponse(
                status=400,
                data={
                    "status": "error",
                    "message": "There is no current system id present",
                },
            )


class EditorAutocomplete(View):
    template_name = "controls/editor.html"

    def get(self, request):
        """Add an existing element and its statements to a system"""

        if request.method != "GET":
            return HttpResponseNotAllowed(["GET"])

        form_dict = dict(request.GET)
        form_values = {}
        for key in form_dict.keys():
            form_values[key] = form_dict[key][0]
        # Form values from ajax data

        if "system_id" in form_values.keys():
            system_id = form_values["system_id"]

            # Does user have permission to add element?
            # Check user permissions
            system = System.objects.get(pk=system_id)

            if not request.user.has_perm("change_system", system):
                # User does not have write permissions
                # Log permission to save answer denied
                logger.info(
                    event="change_system permission_denied",
                    object={
                        "object": "element",
                        "producer_element_name": form_values["producer_element_name"],
                    },
                    user={"id": request.user.id, "username": request.user.username},
                )
                return HttpResponseForbidden(
                    f"Permission denied. {request.user.username} does not have change \
                        privileges to system and/or project."
                )

            selected_controls = system.root_element.controls.all()
            selected_controls_ids = set()
            for sc in selected_controls:
                selected_controls_ids.add(
                    "{} {}".format(sc.oscal_ctl_id, sc.oscal_catalog_key)
                )

            # Add element
            # Look up the element

            text = form_values["text"]

            # The final elements that are returned to the new dropdown created...
            producer_system_elements = Element.objects.filter(
                element_type="system_element"
            ).filter(name__contains=text)

            producer_elements = [
                {"id": str(ele.id), "name": ele.name}
                for ele in producer_system_elements
            ]

            results = {"producer_element_name_value": producer_elements}
            data = json.dumps(results)
            mimetype = "application/json"
            if data:
                return HttpResponse(data, mimetype)
            else:
                return JsonResponse(
                    status=400,
                    data={
                        "status": "error",
                        "message": f"No statements found with the search: {text}",
                    },
                )
        else:
            return JsonResponse(
                status=400,
                data={
                    "status": "error",
                    "message": "There is no current system id present",
                },
            )

    def post(self, request, system_id):
        """Add an existing element and its statements to a system"""
        if request.method != "POST":
            return HttpResponseNotAllowed(["POST"])

        form_dict = dict(request.POST)

        form_values = {}
        for key in form_dict.keys():
            if key == "relatedcomps":
                form_values[key] = form_dict[key]
            else:
                form_values[key] = form_dict[key][0]
        # Form values from ajax data

        if "system_id" in form_values.keys():
            system_id = form_values["system_id"]
            # Does user have permission to add element?
            # Check user permissions
            system = System.objects.get(pk=system_id)
            if not request.user.has_perm("change_system", system):
                # User does not have write permissions
                # Log permission to save answer denied
                logger.info(
                    event="change_system permission_denied",
                    object={
                        "object": "element",
                        "entered_producer_element_name": form_values["text"] or "None",
                    },
                    user={"id": request.user.id, "username": request.user.username},
                )
                return HttpResponseForbidden(
                    f"Permission denied. {request.user.username} does not have change \
                        privileges to system and/or project."
                )

            selected_controls = system.root_element.controls.all()
            selected_controls_ids = set()
            for sc in selected_controls:
                selected_controls_ids.add(
                    "{} {}".format(sc.oscal_ctl_id, sc.oscal_catalog_key)
                )

            # Add element
            if form_values.get("relatedcomps", ""):

                for related_element in form_values["relatedcomps"]:

                    # Look up the element
                    for smt in Statement.objects.filter(
                        id=related_element,
                        statement_type=StatementTypeEnum.CONTROL_IMPLEMENTATION.name,
                    ):
                        logger.info(
                            f"Adding an element with the id {smt.id} and sid class {smt.sid} \
                                to system_id {system_id}"
                        )
                        # Only add statements for controls selected for system
                        if (
                            "{} {}".format(smt.sid, smt.sid_class)
                            in selected_controls_ids
                        ):
                            logger.info(f"smt {smt}")
                            smt.create_system_control_smt_from_component_prototype_smt(
                                system.root_element.id
                            )
                        else:
                            logger.error(
                                f"not adding smt from selected controls for the current \
                                    system: {smt}"
                            )

        # Redirect to the page where the component was added from
        return HttpResponseRedirect(request.META.get("HTTP_REFERER", "/"))


# Baselines
@login_required
def assign_baseline(request, system_id, catalog_key, baseline_name):
    """
    Assign a baseline to a system root element thereby showing selected controls for the system.
    """

    system = System.objects.get(pk=system_id)
    # system.root_element.assign_baseline_controls(user, 'NIST_SP-800-53_rev4', 'low')
    assign_results = system.root_element.assign_baseline_controls(
        request.user, catalog_key, baseline_name
    )
    if assign_results:
        messages.add_message(
            request,
            messages.INFO,
            f'Control baseline {catalog_key.replace("_", " ")} {baseline_name.title()} \
                assigned.',
        )
        # Log start app / new project
        logger.info(
            event="assign_baseline",
            object={
                "object": "system",
                "id": system.root_element.id,
                "title": system.root_element.name,
            },
            baseline={"catalog_key": catalog_key, "baseline_name": baseline_name},
            user={"id": request.user.id, "username": request.user.username},
        )
    else:
        messages.add_message(
            request,
            messages.ERROR,
            f'Control baseline {catalog_key.replace("_", " ")} {baseline_name.title()} \
                assignment failed.',
        )

    return HttpResponseRedirect(f"/systems/{system_id}/controls/selected")


# Export OpenControl
@login_required
def export_system_opencontrol(request, system_id):
    """Export entire system in OpenControl"""

    # Does user have permission
    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("view_system", system):
        # Retrieve primary system Project

        # Create temporary directory structure
        import tempfile

        temp_dir = tempfile.TemporaryDirectory(dir=".")
        repo_path = os.path.join(
            temp_dir.name, system.root_element.name.replace(" ", "_")
        )
        if not os.path.exists(repo_path):
            os.makedirs(repo_path)

        # Create various directories
        os.makedirs(os.path.join(repo_path, "components"))
        os.makedirs(os.path.join(repo_path, "standards"))
        os.makedirs(os.path.join(repo_path, "certifications"))

        # Create opencontrol.yaml config file
        cfg_str = """schema_version: 1.0.0
name: ~
metadata:
  authorization_id: ~
  description: ~
  organization:
    name: ~
    abbreviation: ~
  repository: ~
components: []
standards:
- ./standards/NIST-SP-800-53-rev4.yaml
certifications:
- ./certifications/fisma-low-impact.yaml
"""

        # read default opencontrol.yaml into object
        cfg = rtyaml.load(cfg_str)
        # customize values
        cfg["name"] = system.root_element.name

        with open(os.path.join(repo_path, "opencontrol.yaml"), "w") as outfile:
            outfile.write(rtyaml.dump(cfg))

        # Populate reference directories from reference
        OPENCONTROL_PATH = os.path.join(
            os.path.dirname(__file__), "data", "opencontrol"
        )
        shutil.copyfile(
            os.path.join(OPENCONTROL_PATH, "standards", "NIST-SP-800-53-rev4.yaml"),
            os.path.join(repo_path, "standards", "NIST-SP-800-53-rev4.yaml"),
        )
        shutil.copyfile(
            os.path.join(OPENCONTROL_PATH, "standards", "NIST-SP-800-171r1.yaml"),
            os.path.join(repo_path, "standards", "NIST-SP-800-53-rev4.yaml"),
        )
        shutil.copyfile(
            os.path.join(OPENCONTROL_PATH, "standards", "opencontrol.yaml"),
            os.path.join(repo_path, "standards", "opencontrol.yaml"),
        )
        shutil.copyfile(
            os.path.join(OPENCONTROL_PATH, "standards", "hipaa-draft.yaml"),
            os.path.join(repo_path, "standards", "hipaa-draft.yaml"),
        )
        shutil.copyfile(
            os.path.join(OPENCONTROL_PATH, "certifications", "fisma-low-impact.yaml"),
            os.path.join(repo_path, "certifications", "fisma-low-impact.yaml"),
        )

        # Populate system information files

        # Populate component files
        if not os.path.exists(os.path.join(repo_path, "components")):
            os.makedirs(os.path.join(repo_path, "components"))
        for element in system.producer_elements:
            # Build OpenControl
            ocf = {
                "name": element.name,
                "schema_version": "3.0.0",
                "documentation_complete": False,
                "satisfies": [],
            }
            satisfies_smts = ocf["satisfies"]
            # Retrieve impl_smts produced by element and consumed by system
            # Get the impl_smts contributed by this component to system
            impl_smts = element.statements_produced.filter(
                consumer_element=system.root_element
            )
            for smt in impl_smts:
                my_dict = {
                    "control_key": smt.sid.upper(),
                    "control_name": smt.catalog_control_as_dict["title"],
                    "standard_key": smt.sid_class,
                    "covered_by": [],
                    "security_control_type": "Hybrid | Inherited | ...",
                    "narrative": [{"text": smt.body}],
                    "remarks": [{"text": smt.remarks}],
                }
                satisfies_smts.append(my_dict)
            opencontrol_string = rtyaml.dump(ocf)
            # Write component file
            with open(
                os.path.join(
                    repo_path,
                    "components",
                    "{}.yaml".format(element.name.replace(" ", "_")),
                ),
                "w",
            ) as fh:
                fh.write(opencontrol_string)

        # Build Zip archive
        # TODO Make Temporary File
        #      Current approach leads to race conditions!
        shutil.make_archive("/tmp/Zipped_file", "zip", repo_path)

        # Download Zip archive of OpenControl files
        with open("/tmp/Zipped_file.zip", "rb") as tmp:
            tmp.seek(0)
            stream = tmp.read()
            blob = stream
        mime_type = "application/octet-stream"
        filename = "{}-opencontrol-{}.zip".format(
            system.root_element.name.replace(" ", "_"),
            datetime.now().strftime("%Y-%m-%d-%H-%M"),
        )

        resp = HttpResponse(blob, mime_type)
        resp["Content-Disposition"] = "inline; filename=" + filename

        # Clean up
        shutil.rmtree(repo_path)
        # os.remove("/tmp/Zipped_file") ????
        return resp

    else:
        # User does not have permission to this system
        raise Http404


# PoamS
@login_required
def poams_list(request, system_id):
    """List Poams for a system"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("view_system", system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        controls = system.root_element.controls.all()
        poam_smts = system.root_element.statements_consumed.filter(
            statement_type="POAM"
        ).order_by("-updated")

        # Return the controls
        context = {
            "system": system,
            "project": project,
            "controls": controls,
            "poam_smts": poam_smts,
            "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
        }
        return render(request, "systems/poams_list.html", context)
    else:
        # User does not have permission to this system
        raise Http404


@login_required
def new_poam(request, system_id):
    """Form to create new POAM"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("view_system", system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        controls = system.root_element.controls.all()

        if request.method == "POST":
            statement_form = StatementPoamForm(request.POST)
            # if statement_form.is_valid() and poam_form.is_valid():
            if statement_form.is_valid():
                statement = statement_form.save()
                poam_form = PoamForm(request.POST)
                if poam_form.is_valid():
                    poam = poam_form.save()
                    print("POAM ID", poam.get_next_poam_id(system))
                    poam.poam_id = poam.get_next_poam_id(system)
                    poam.statement = statement
                    poam.save()
                return HttpResponseRedirect("/systems/{}/poams".format(system_id), {})
                # url(r'^(?P<system_id>.*)/poams$', views.poams_list, name="poams_list"),
            else:
                pass
        else:
            statement_form = StatementPoamForm(
                status="Open",
                statement_type="POAM",
                consumer_element=system.root_element,
            )
            poam = Poam()
            poam_form = PoamForm()
            return render(
                request,
                "systems/poam_form.html",
                {
                    "statement_form": statement_form,
                    "poam_form": poam_form,
                    "system": system,
                    "project": project,
                    "controls": controls,
                },
            )
    else:
        # User does not have permission to this system
        raise Http404


@login_required
def edit_poam(request, system_id, poam_id):
    """Form to create new POAM"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("view_system", system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        controls = system.root_element.controls.all()
        # Retrieve POAM Statement
        poam_smt = get_object_or_404(Statement, id=poam_id)

        if request.method == "POST":
            statement_form = StatementPoamForm(request.POST, instance=poam_smt)
            poam_form = PoamForm(request.POST, instance=poam_smt.poam)
            if statement_form.is_valid() and poam_form.is_valid():
                # Save statement after updating values
                statement_form.save()
                poam_form.save()
                return HttpResponseRedirect("/systems/{}/poams".format(system_id), {})
            else:
                pass
                # TODO: What if form invalid?
        else:
            statement_form = StatementPoamForm(
                initial={
                    "statement_type": poam_smt.statement_type,
                    "status": poam_smt.status,
                    "consumer_element": system.root_element,
                    "body": poam_smt.body,
                    "remarks": poam_smt.remarks,
                }
            )
            poam_form = PoamForm(
                initial={
                    "weakness_name": poam_smt.poam.weakness_name,
                    "controls": poam_smt.poam.controls,
                    "poam_group": poam_smt.poam.poam_group,
                    "risk_rating_original": poam_smt.poam.risk_rating_original,
                    "risk_rating_adjusted": poam_smt.poam.risk_rating_adjusted,
                    "weakness_detection_source": poam_smt.poam.weakness_detection_source,
                    "remediation_plan": poam_smt.poam.remediation_plan,
                    "milestones": poam_smt.poam.milestones,
                    "scheduled_completion_date": poam_smt.poam.scheduled_completion_date,
                }
            )
            return render(
                request,
                "systems/poam_edit_form.html",
                {
                    "statement_form": statement_form,
                    "poam_form": poam_form,
                    "system": system,
                    "project": project,
                    "controls": controls,
                    "poam_smt": poam_smt,
                },
            )
    else:
        # User does not have permission to this system
        raise Http404


@login_required
def poam_export_xlsx(request, system_id):
    return poam_export(request, system_id, "xlsx")


@login_required
def poam_export_csv(request, system_id):
    return poam_export(request, system_id, "csv")


@login_required
def poam_export(request, system_id, format="xlsx"):
    """Export POA&M in either xlsx or csv"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected POA&Ms if user has permission on system
    if request.user.has_perm("view_system", system):

        if format == "xlsx":
            from tempfile import NamedTemporaryFile

            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "POA&Ms"
        else:
            import csv
            import io

            csv_buffer = io.StringIO(newline="\n")
            csv_writer = csv.writer(csv_buffer)

        poam_fields = [
            {"var_name": "poam_id", "name": "POA&M ID", "width": 8},
            {"var_name": "poam_group", "name": "POA&M Group", "width": 16},
            {"var_name": "weakness_name", "name": "Weakness Name", "width": 24},
            {"var_name": "controls", "name": "Controls", "width": 16},
            {"var_name": "body", "name": "Description", "width": 60},
            {"var_name": "status", "name": "Status", "width": 8},
            {
                "var_name": "risk_rating_original",
                "name": "Risk Rating Original",
                "width": 16,
            },
            {
                "var_name": "risk_rating_adjusted",
                "name": "Risk Rating Adjusted",
                "width": 16,
            },
            {
                "var_name": "weakness_detection_source",
                "name": "Weakness Detection Source",
                "width": 24,
            },
            {
                "var_name": "weakness_source_identifier",
                "name": "Weakness Source Identifier",
                "width": 24,
            },
            {"var_name": "remediation_plan", "name": "Remediation Plan", "width": 60},
            {"var_name": "milestones", "name": "Milestones", "width": 60},
            {"var_name": "milestone_changes", "name": "Milestone Changes", "width": 30},
            {
                "var_name": "scheduled_completion_date",
                "name": "Scheduled Completion Date",
                "width": 18,
            },
        ]

        # create header row
        column = 0
        ord_zeroth_column = ord("A") - 1
        csv_row = []

        for poam_field in poam_fields:
            column += 1
            if format == "xlsx":
                c = ws.cell(row=1, column=column, value=poam_field["name"])
                c.fill = PatternFill("solid", fgColor="5599FE")
                c.font = Font(color="FFFFFF", bold=True)
                c.border = Border(
                    left=Side(border_style="thin", color="444444"),
                    right=Side(border_style="thin", color="444444"),
                    bottom=Side(border_style="thin", color="444444"),
                    outline=Side(border_style="thin", color="444444"),
                )
                ws.column_dimensions[
                    chr(ord_zeroth_column + column)
                ].width = poam_field["width"]
            else:
                csv_row.append(poam_field["name"])
        # Add column for URL
        if format == "xlsx":
            c = ws.cell(row=1, column=column, value="URL")
            c.fill = PatternFill("solid", fgColor="5599FE")
            c.font = Font(color="FFFFFF", bold=True)
            c.border = Border(
                left=Side(border_style="thin", color="444444"),
                right=Side(border_style="thin", color="444444"),
                bottom=Side(border_style="thin", color="444444"),
                outline=Side(border_style="thin", color="444444"),
            )
            ws.column_dimensions[chr(ord_zeroth_column + column)].width = 60
        else:
            csv_row.append("URL")

        if format != "xlsx":
            csv_writer.writerow(csv_row)

        # Retrieve POA&Ms and create POA&M rows
        poam_smts = system.root_element.statements_consumed.filter(
            statement_type="POAM"
        ).order_by("id")
        row = 1
        for poam_smt in poam_smts:
            csv_row = []
            row += 1

            # Loop through fields
            column = 0
            for poam_field in poam_fields:
                column += 1
                if format == "xlsx":
                    if poam_field["var_name"] in ["body", "status"]:
                        c = ws.cell(
                            row=row,
                            column=column,
                            value=getattr(poam_smt, poam_field["var_name"]),
                        )
                    else:
                        if poam_field["var_name"] == "poam_id":
                            c = ws.cell(
                                row=row,
                                column=column,
                                value="V-{}".format(
                                    getattr(poam_smt.poam, poam_field["var_name"])
                                ),
                            )
                        else:
                            c = ws.cell(
                                row=row,
                                column=column,
                                value=getattr(poam_smt.poam, poam_field["var_name"]),
                            )
                    c.fill = PatternFill("solid", fgColor="FFFFFF")
                    c.alignment = Alignment(
                        vertical="top", horizontal="left", wrapText=True
                    )
                    c.border = Border(
                        right=Side(border_style="thin", color="444444"),
                        bottom=Side(border_style="thin", color="444444"),
                        outline=Side(border_style="thin", color="444444"),
                    )
                else:
                    if poam_field["var_name"] in ["body", "status"]:
                        csv_row.append(getattr(poam_smt, poam_field["var_name"]))
                    else:
                        if poam_field["var_name"] == "poam_id":
                            csv_row.append(
                                "V-{}".format(
                                    getattr(poam_smt.poam, poam_field["var_name"])
                                )
                            )
                        else:
                            csv_row.append(
                                getattr(poam_smt.poam, poam_field["var_name"])
                            )

            # Add URL column
            poam_url = settings.SITE_ROOT_URL + "/systems/{}/poams/{}/edit".format(
                system_id, poam_smt.id
            )
            if format == "xlsx":
                c = ws.cell(row=row, column=column, value=poam_url)
                c.fill = PatternFill("solid", fgColor="FFFFFF")
                c.alignment = Alignment(
                    vertical="top", horizontal="left", wrapText=True
                )
                c.border = Border(
                    right=Side(border_style="thin", color="444444"),
                    bottom=Side(border_style="thin", color="444444"),
                    outline=Side(border_style="thin", color="444444"),
                )
            else:
                csv_row.append(poam_url)

            if format != "xlsx":
                csv_writer.writerow(csv_row)

        if format == "xlsx":
            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                tmp.seek(0)
                stream = tmp.read()
                blob = stream
        else:
            blob = csv_buffer.getvalue()
            csv_buffer.close()

        # Determine filename based on system name
        system_name = system.root_element.name.replace(" ", "_") + "_" + system_id
        filename = "{}_poam_export-{}.{}".format(
            system_name, datetime.now().strftime("%Y-%m-%d-%H-%M"), format
        )
        mime_type = "application/octet-stream"

        resp = HttpResponse(blob, mime_type)
        resp["Content-Disposition"] = "inline; filename=" + filename
        return resp
    else:
        # User does not have permission to this system
        raise Http404


# Project
@login_required
def project_import(request, project_id):
    """
    Import an entire project's components and control content
    """
    project = Project.objects.get(id=project_id)
    system_id = project.system.id
    # Retrieve identified System
    system = System.objects.get(id=system_id)
    system_root_element = system.root_element

    # Retrieve identified System
    if request.method == "POST":
        project_data = request.POST["json_content"]
        # Need to get or create the app source by the id of the given app source
        # module_name = json.loads(project_data).get('project').get('module').get('key')
        title = json.loads(project_data).get("project").get("title")
        system_root_element.name = title

        logger.info(
            event="project JSON import update",
            object={"object": "project", "id": project.id, "title": project.title},
            user={"id": request.user.id, "username": request.user.username},
        )
        messages.add_message(
            request, messages.INFO, f"Project {project.id} was updated."
        )

        # Import questionnaire data
        log_output = []
        try:
            from collections import OrderedDict

            data = json.loads(project_data, object_pairs_hook=OrderedDict)
        except Exception:
            log_output.append("There was an error reading the export file.")
        else:
            try:
                # Update project data.
                project.import_json(
                    data, request.user, "imp", lambda x: log_output.append(x)
                )
            except Exception as e:
                log_output.append(str(e))

        # Log output
        logger.info(
            event="project JSON import",
            object={
                "object": "project",
                "id": project.id,
                "title": project.title,
                "log_output": log_output,
            },
            user={"id": request.user.id, "username": request.user.username},
        )

        # Import components and their statements
        loaded_imported_jsondata = json.loads(project_data)
        if loaded_imported_jsondata.get("component-definitions") is not None:
            # Load and get the components then dump
            comp_num = 0
            for k, val in enumerate(
                loaded_imported_jsondata.get("component-definitions")
            ):
                oscal_component_json = json.dumps(
                    loaded_imported_jsondata.get("component-definitions")[k]
                )
                import_name = request.POST.get("import_name", "")
                import_record = ComponentImporter().import_components_as_json(
                    import_name, oscal_component_json, request
                )
                if import_record is not None:
                    comps = add_selected_components(system, import_record)
                    comp_num = comp_num + len(comps)
            messages.add_message(
                request, messages.INFO, f"{comp_num} components were created."
            )

            # Import Poams
        if loaded_imported_jsondata.get("poams") is not None:
            # Load and get the poams then dump
            poam_num = 0
            for k, poam in enumerate(loaded_imported_jsondata.get("poams")):
                # Create a Poam for the system
                # Statement linked to the poam
                poamsmt_data = poam.get("statement")
                poam_smt = Statement.objects.create(
                    sid=None,
                    sid_class=None,
                    pid=None,
                    body=poamsmt_data.get("body"),
                    remarks=poam.get("remarks"),
                    version=poam.get("version"),
                    statement_type="POAM",
                    status=poamsmt_data.get("status", "New"),
                    uuid=str(uuid4()),
                    consumer_element=system_root_element,
                )
                # Create Poam with statement and imported data
                poam = Poam.objects.create(
                    statement=poam_smt,
                    controls=poam.get("controls"),
                    milestones=poam.get("milestones"),
                    poam_id=Poam.objects.order_by("-poam_id")[0] + 1
                    if poam.get("poam_id") is None
                    else poam.get("poam_id"),
                    remediation_plan=poam.get("remediation_plan"),
                    risk_rating_adjusted=poam.get("risk_rating_adjusted"),
                    risk_rating_original=poam.get("risk_rating_original"),
                    scheduled_completion_date=poam.get("scheduled_completion_date"),
                    weakness_detection_source=poam.get("weakness_detection_source"),
                    weakness_name=poam.get("weakness_name"),
                    weakness_source_identifier=poam.get("weakness_source_identifier"),
                    poam_group=poam.get("poam_group"),
                )
                poam_num += 1
                logger.info(
                    event="Poam import",
                    object={
                        "object": "poam",
                        "id": poam.poam_id,
                        "controls": poam.poam_group,
                    },
                    user={"id": request.user.id, "username": request.user.username},
                )
            messages.add_message(
                request, messages.INFO, f"{poam_num} POA&Ms were created."
            )

        return HttpResponseRedirect("/projects")


@login_required
def project_export(request, project_id):
    """
    Export an entire project's components and control content
    """
    # Of the project in the current system. pick one project to export
    project = Project.objects.get(id=project_id)
    system_id = project.system.id
    # Retrieve identified System
    system = System.objects.get(id=system_id)
    system_root_element = system.root_element

    # Retrieve related selected controls and Poams if user has permission on system
    if request.user.has_perm("view_system", system):
        oscal_comps = []
        for element in system.producer_elements:
            # Implementation statement OSCAL JSON
            impl_smts = element.statements_produced.filter(
                consumer_element=system_root_element
            )
            component = OSCALComponentSerializer(element, impl_smts).as_json()
            oscal_comps.append(component)
        poams = []
        poam_smts = system_root_element.statements_consumed.filter(
            statement_type="POAM"
        ).order_by("id")
        for smt in poam_smts:
            poam = {
                "controls": smt.poam.controls,
                "milestones": smt.poam.milestones,
                "poam_id": smt.poam.poam_id,
                "remediation_plan": smt.poam.remediation_plan,
                "risk_rating_original": smt.poam.risk_rating_original,
                "risk_rating_adjusted": smt.poam.risk_rating_adjusted,
                "scheduled_completion_date": smt.poam.scheduled_completion_date,
                "weakness_detection_source": smt.poam.weakness_detection_source,
                "weakness_name": smt.poam.weakness_name,
                "weakness_source_identifier": smt.poam.weakness_source_identifier,
                "poam_group": smt.poam.poam_group,
                "statement": {
                    "body": smt.body,
                    "remarks": smt.remarks,
                    "version": smt.remarks,
                    "status": smt.status,
                    "uuid": str(smt.uuid),
                },
            }
            # Add json version as an element in the poams list
            poams.append(json.dumps(poam))
    questionnaire_data = json.dumps(
        project.export_json(include_metadata=True, include_file_content=True)
    )
    data = json.loads(questionnaire_data)
    data["component-definitions"] = [
        json.loads(oscal_comp) for oscal_comp in oscal_comps
    ]
    data["poams"] = [json.loads(poam) for poam in poams]
    response = JsonResponse(data, json_dumps_params={"indent": 2})
    filename = (
        project.title.replace(" ", "_")
        + "-"
        + datetime.now().strftime("%Y-%m-%d-%H-%M")
    )
    response["Content-Disposition"] = f'attachment; filename="{quote(filename)}.json"'
    return response


# System OSCAL
@login_required
def system_profile_oscal_json(request, system_id):
    """
    Return an OSCAL profile for this system.
    TODO: for now, we return an empty response.
    """

    data = {}
    return JsonResponse(data)


@login_required
def system_deployments(request, system_id):
    """List deployments for a system"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("view_system", system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]

        # Retrieve list of deployments for the system
        deployments = system.deployments.all().order_by(Lower("name"))

        # Return the controls
        context = {
            "system": system,
            "project": project,
            "deployments": deployments,
        }
        return render(request, "systems/deployments_list.html", context)
    else:
        # User does not have permission to this system
        raise Http404


@login_required
def manage_system_deployment(request, system_id, deployment_id=None):
    """Form to create or edit system deployment"""

    # Can user view this sytem?
    system = System.objects.get(id=system_id)
    if not request.user.has_perm("view_system", system):
        # User does not have permission to this system
        raise Http404

    di = get_object_or_404(Deployment, pk=deployment_id) if deployment_id else None
    if request.method == "POST":
        form = DeploymentForm(request.POST, instance=di)
        if form.is_valid():
            form.save()
            deployment = form.instance
            # Create message to display to user
            if di:
                messages.add_message(
                    request, messages.INFO, f"{deployment.name} deployment was edited."
                )
                logger.info(
                    event="edit_deployment",
                    object={
                        "object": "deployment",
                        "id": deployment.id,
                        "name": deployment.name,
                    },
                    user={"id": request.user.id, "username": request.user.username},
                )
            else:
                messages.add_message(
                    request, messages.INFO, f"{deployment.name} deployment was created."
                )
                logger.info(
                    event="create_deployment",
                    object={
                        "object": "deployment",
                        "id": deployment.id,
                        "name": deployment.name,
                    },
                    user={"id": request.user.id, "username": request.user.username},
                )
            return redirect("system_deployments", system_id=system_id)
    else:
        if di is None:
            di = Deployment(system_id=system_id)
        form = DeploymentForm(instance=di)

    return render(
        request,
        "systems/deployment_form.html",
        {
            "form": form,
            "deployment": di,
        },
    )


@login_required
def deployment_history(request, system_id, deployment_id=None):
    """Returns the history for the given deployment"""

    # Can user view this sytem?
    system = System.objects.get(id=system_id)
    if not request.user.has_perm("view_system", system):
        # User does not have permission to this system
        raise Http404

    from controls.models import Deployment

    full_dpt_history = None
    try:
        deployments = Deployment.objects.get(id=deployment_id)
        full_dpt_history = deployments.history.all()
    except Deployment.DoesNotExist:
        messages.add_message(
            request,
            messages.ERROR,
            "The deployment ID is not valid. Please check if this deployment is still in \
                Blueprint.",
        )
    context = {
        "deployment": full_dpt_history,
    }
    return render(request, "systems/deployment_history.html", context)


@login_required
def system_deployment_inventory(request, system_id, deployment_id):
    """List system deployment inventory"""

    # Retrieve identified System
    system = System.objects.get(id=system_id)
    # Retrieve related selected controls if user has permission on system
    if request.user.has_perm("view_system", system):
        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]

        deployment = get_object_or_404(Deployment, pk=deployment_id)

        inventory_items = (
            [item for item in deployment.inventory_items]
            if deployment.inventory_items is not None
            else []
        )

        # Return the controls
        context = {
            "system": system,
            "project": project,
            "deployment": deployment,
            "inventory_items": inventory_items,
            # "controls": controls,
            # "poam_smts": poam_smts,
            # "enable_experimental_opencontrol": SystemSettings.enable_experimental_opencontrol,
            # "enable_experimental_oscal": SystemSettings.enable_experimental_oscal,
        }
        return render(request, "systems/deployment_inventory.html", context)
    else:
        # User does not have permission to this system
        raise Http404


@login_required
def system_assessment_results_list(request, system_id=None):
    """List System Assessment Results for a system"""

    # Retrieve identified System
    if system_id:
        # Can user view this sytem?
        system = System.objects.get(id=system_id)
        if not request.user.has_perm("view_system", system):
            # User does not have permission to this system
            raise Http404

        system = System.objects.get(id=system_id)
        # Retrieve related selected controls if user has permission on system
        if not request.user.has_perm("view_system", system):
            # User does not have permission to this system
            raise Http404

        # Retrieve primary system Project
        # Temporarily assume only one project and get first project
        project = system.projects.all()[0]
        sars = system.system_assessment_result.all().order_by("created").reverse()

        # Retrieve user's API keys
        api_keys = request.user.get_api_keys()

        context = {
            "system": system,
            "project": project,
            "sars": sars,
            "api_key_ro": api_keys["ro"],
            "api_key_rw": api_keys["rw"],
            "api_key_wo": api_keys["wo"],
        }
        return render(request, "systems/sar_list.html", context)


@login_required
def view_system_assessment_result_summary(request, system_id, sar_id=None):
    """View Summary of System Assessment Results"""

    # Can user view this sytem?
    system = System.objects.get(id=system_id)
    if not request.user.has_perm("view_system", system):
        # User does not have permission to this system
        raise Http404

    # Retrieve primary system Project
    # Temporarily assume only one project and get first project
    project = system.projects.all()[0]
    sar = get_object_or_404(SystemAssessmentResult, pk=sar_id) if sar_id else None

    # Get assessment targets results from wrapped SAR data
    sar_items = (
        [item for item in sar.assessment_results["sar"]]
        if sar.assessment_results is not None
        else []
    )

    # Get summary pass fail across all assessment results included collection
    # TODO: note high/low category
    summary = {}
    for param in ["pass", "fail", "other", "unknown", "error"]:
        summary[param] = sum(d[param] for d in sar_items if d and param in d)

    return render(
        request,
        "systems/sar_summary.html",
        {
            "project": project,
            "sar": sar,
            "sar_items": sar_items,
            "assessment_results_json": json.dumps(
                sar.assessment_results, indent=4, sort_keys=True
            ),
            "summary": summary,
        },
    )


@login_required
def manage_system_assessment_result(request, system_id, sar_id=None):
    """Form to create or edit system assessment result"""

    # Can user view this system?
    system = System.objects.get(id=system_id)
    if not request.user.has_perm("view_system", system):
        # User does not have permission to this system
        raise Http404

    sari = get_object_or_404(SystemAssessmentResult, pk=sar_id) if sar_id else None
    if request.method == "POST":
        form = SystemAssessmentResultForm(request.POST, instance=sari)
        if form.is_valid():
            form.save()
            sar = form.instance
            # Create message to display to user
            if sari:
                messages.add_message(
                    request,
                    messages.INFO,
                    f"System assessment result {sar.name} was edited.",
                )
                logger.info(
                    event="edit_system_assessment_result",
                    object={
                        "object": "system_assessment_result",
                        "id": sar.id,
                        "name": sar.name,
                    },
                    user={"id": request.user.id, "username": request.user.username},
                )
            else:
                messages.add_message(
                    request,
                    messages.INFO,
                    f"System assessment result {sar.name} was created.",
                )
                logger.info(
                    event="create_system_assessment_result",
                    object={
                        "object": "system_assessment_result",
                        "id": sar.id,
                        "name": sar.name,
                    },
                    user={"id": request.user.id, "username": request.user.username},
                )
            return redirect("system_assessment_results_list", system_id=system_id)
    else:
        if sari is None:
            sari = SystemAssessmentResult(system_id=system_id)
        form = SystemAssessmentResultForm(instance=sari)
        # Filter deployments to current system
        form.fields["deployment"].queryset = Deployment.objects.filter(
            system__id=system_id
        )

    return render(
        request,
        "systems/sar_form.html",
        {
            "form": form,
            "system_id": system_id,
        },
    )


@login_required
def system_assessment_result_history(request, system_id, sar_id=None):
    """Returns the history for the given deployment system assessment result"""

    # TODO check user permission to view
    full_sar_history = None
    try:
        sar = SystemAssessmentResult.objects.get(id=sar_id)
        full_sar_history = sar.history.all()
    except SystemAssessmentResult.DoesNotExist:
        messages.add_message(
            request,
            messages.ERROR,
            "The system assessment result ID is not valid. Please check if this system \
                assessment result is still in Blueprint.",
        )
    context = {
        "deployment": full_sar_history,
    }
    return render(request, "systems/sar_history.html", context)


@login_required
def new_system_assessment_result_wazuh(request, system_id):
    """Returns a SAR info from Wazuh and adds to system"""

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    else:
        # Validate data
        valid = True
        for param in ["wazuhhost_val", "user_val", "passwd_val", "agents_val"]:
            if param not in request.POST or request.POST[param] == "":
                valid = False
                messages.add_message(
                    request,
                    messages.WARNING,
                    f'Please complete the field {param.replace("_val","")}',
                )
        if not valid:
            return HttpResponseRedirect(f"/systems/{system_id}/assessments")

        # Check user permissions
        system = System.objects.get(pk=system_id)
        if not request.user.has_perm("change_system", system):
            # User does not have write permissions
            return HttpResponseForbidden(
                f"Permission denied. {request.user.username} does not have change privileges \
                    to system and/or project."
            )

        from sec_srvc.wazuh import WazuhSecurityService

        wazuh_sec_svc = WazuhSecurityService()
        wazuh_sec_svc.setup(base_url=request.POST["wazuhhost_val"])

        authentication = wazuh_sec_svc.authenticate(
            request.POST["user_val"], request.POST["passwd_val"]
        )

        if wazuh_sec_svc.is_authenticated:
            identifiers = request.POST["agents_val"]
            extracted_data = wazuh_sec_svc.extract_data(authentication, identifiers)

            # TODO: Set deployment id
            deployment_uuid = None

            transformed_data = wazuh_sec_svc.transform_data(
                extracted_data,
                system_id,
                "Scan Title",
                "Scan description",
                deployment_uuid,
            )

            # Determine deployment_id from deployment_uuid
            # TODO: Make sure deployment is associated with system
            if deployment_uuid is None or deployment_uuid == "None":
                # When deployment is not defined, leave blank and attach SAR to system only
                deployment = None
                deployment_id = None
            else:
                deployment = Deployment.objects.get(uuid=deployment_uuid)
                deployment_id = deployment.id

            sar = SystemAssessmentResult(
                name=transformed_data["metadata"]["title"],
                description=transformed_data["metadata"]["description"],
                system_id=transformed_data["metadata"]["system_id"],
                deployment_id=deployment_id,
                assessment_results=transformed_data,
            )
            sar.save()
            logger.info(
                event="create_system_assessment_result",
                object={
                    "object": "system_assessment_result",
                    "id": sar.id,
                    "name": sar.name,
                },
                user={"id": request.user.id, "username": request.user.username},
            )
            messages.add_message(
                request, messages.INFO, "Data from Wazuh was retrieved and loaded"
            )

            # Redirect
            return HttpResponseRedirect(f"/systems/{system_id}/assessments")

        else:
            # TODO: better handling of response code; 401, 301, etc.
            raise Exception(wazuh_sec_svc.error_msg["error"])


@login_required
def update_control_status(request, system_id):
    """Update the control status"""
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    else:
        form_dict = dict(request.POST)

        eid = form_dict["eid"][0]
        status = form_dict["status"][0]

        control = ElementControl.objects.get(id=eid)
        control.status = status

        system = System.objects.get(id=system_id)

        if not request.user.has_perm("change_system", system):
            # User does not have write permissions
            # Log permission to save answer denied
            logger.info(
                event="update_control_status permission_denied",
                object={"object": "control", "id": control.id},
                user={"id": request.user.id, "username": request.user.username},
            )
            return HttpResponseForbidden(
                "Permission denied. {} does not have change privileges to system and/or \
                    project.".format(
                    request.user.username
                )
            )

        if control is None:
            control_msg = "The id for this control is no longer valid in the database."
            return JsonResponse({"message": control_msg}, status=422)

        if status not in set(
            str(code) for (code, label) in ElementControl.Statuses.choices
        ):
            control_msg = "Status choice not allowed."
            return JsonResponse({"message": control_msg}, status=422)

        try:
            control.save()
            control_status = 200
            control_msg = "Controls status updated;"
        except Exception as e:
            control_msg = "Control status update failed. Error reported {}".format(e)
            return JsonResponse({"message": control_msg}, status=422)

        return JsonResponse({"message": control_msg}, status=control_status)
